import re
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.db import IntegrityError, connection
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.http import JsonResponse
User = get_user_model()
from . import utils
from datetime import datetime
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter
from decimal import Decimal


# Funcionalidad principal

def clean_null(value):
    return value if value not in ['', None] else None


def home(request):
    return render(request, 'home.html')


def signup(request):

    if request.method == 'GET':
        return render(request, 'signin.html', {
            'form': UserCreationForm
        })
    else:
        if request.POST['password1'] == request.POST['password2']:
            try:
                user = User.objects.create_user(
                    username=request.POST['username'], password=request.POST['password1'])
                user.save()
                login(request, user)
                return redirect('list_ruta')
            except IntegrityError:
                return render(request, 'signin.html', {
                    'form': UserCreationForm,
                    'error': 'User already exists'
                })
        return render(request, 'signin.html', {
            'form': UserCreationForm,
            'error': 'Password do not match'
        })


def signout(request):
    logout(request)
    return redirect('home')


def signin(request):
    if request.method == 'GET':
        return render(request, 'signin.html', {
            'form': AuthenticationForm
        })
    else:
        user = authenticate(request, username=request.POST.get(
            'username'), password=request.POST.get('password'))

        if user is None:
            return render(request, 'signin.html', {
                'form': AuthenticationForm,
                'error': 'Username or password is incorrect'
            })
        else:
            login(request, user)
            return redirect('list_ruta')
        
# Render de pdf

def render_to_pdf(template_src, context_dict=None):
    """
    Genera un PDF desde una plantilla HTML usando xhtml2pdf (pisa).
    Retorna un HttpResponse con el PDF, o None si hay error.
    """
    if context_dict is None:
        context_dict = {}

    try:
        template = get_template(template_src)
        html = template.render(context_dict)
        result = BytesIO()

        pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

        if not pdf.err:
            return HttpResponse(result.getvalue(), content_type='application/pdf')
        else:
            # Muestra el error en consola para depuración
            print("❌ Error al generar PDF con xhtml2pdf:", pdf.err)
            return None

    except Exception as e:
        # Captura errores inesperados
        print("❌ Excepción en render_to_pdf:", str(e))
        return None


# Funcionalidad de vendedores

@login_required
def list_vendedores(request):
    ide = request.user.ide  # Obtener el 'ide' del usuario autenticado
    try:
        vendedores = utils.get_vendedores(ide)
    except Exception as e:
        messages.error(request, f"Error al cargar los vendedores: {str(e)}")

    return render(request, 'vendedores.html', {'vendedores': vendedores})

@login_required
def view_vendedor(request, idx):
    ide = request.user.ide
    vendedor = None  # Inicializa vendedor como None

    if request.method == 'POST':
        idvend = request.POST.get('idvend')
        nombre = request.POST.get('nombre')
        idalmacen = request.POST.get('idalmacen')
        estado = int(request.POST.get('estado', 0))

        try:
            with connection.cursor() as cursor:
                cursor.execute("EXEC dj_e_vendedor %s, %s, %s, %s, %s, %s", [
                               ide, idvend, nombre, idalmacen, idx, estado])

            messages.success(request, 'Editado correctamente')
            # Redirige después de la edición
            return redirect('list_vendedores')

        except Exception as e:
            error_message = str(e)

            # Filtrar solo el mensaje relevante
            match = re.search(
                r"El vendedor con clave .*? ya existe\.", error_message)
            if match:
                error_message = match.group(0)  # Extrae solo el mensaje limpio

            messages.error(request, error_message)

    else:
        # Obtener datos del vendedor actual
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_e_vendedor %s", [idx])
            row = cursor.fetchone()
            if row:
                vendedor = {
                    'ide': row[0],
                    'idvend': row[1],
                    'nombre': row[2],
                    'idalmacen': row[3],
                    'estado': row[4],
                }

    return render(request, 'vendedores_view.html', {'vendedor': vendedor})


@login_required
def add_vendedor(request):
    ide = request.user.ide  # Obtener el 'ide' del usuario autenticado

    if request.method == 'POST':

        idvend = request.POST.get('idvend')
        nombre = request.POST.get('nombre')
        idalmacen = request.POST.get('idalmacen')
        estado = request.POST.get('estado')

        try:
            with connection.cursor() as cursor:
                cursor.execute("EXEC dj_a_vendedor %s, %s, %s, %s, %s", [
                               ide, idvend, nombre, idalmacen, estado])

            messages.success(request, 'Agregado correctamente')
            # Redirigir si la inserción es exitosa
            return redirect('list_vendedores')

        except Exception as e:
            error_message = str(e)

            # Expresión regular para extraer solo el mensaje de error sin códigos ni detalles técnicos
            match = re.search(
                r"El vendedor con clave .*? ya existe\.", error_message)
            if match:
                error_message = match.group(0)  # Extrae solo el mensaje limpio
            

            messages.error(request, error_message)

    return render(request, 'vendedores_form.html')


@login_required
def edit_vendedor(request, idx):
    ide = request.user.ide
    vendedor = None  # Inicializa vendedor como None

    if request.method == 'POST':
        idvend = request.POST.get('idvend')
        nombre = request.POST.get('nombre')
        idalmacen = request.POST.get('idalmacen')
        estado = request.POST.get('estado')

        try:
            with connection.cursor() as cursor:
                cursor.execute("EXEC dj_e_vendedor %s, %s, %s, %s, %s, %s", [
                               ide, idvend, nombre, idalmacen, idx, estado])

            messages.success(request, 'Editado correctamente')
            # Redirige después de la edición
            return redirect('list_vendedores')

        except Exception as e:
            error_message = str(e)

            # Filtrar solo el mensaje relevante
            match = re.search(
                r"El vendedor con clave .*? ya existe\.", error_message)
            if match:
                error_message = match.group(0)  # Extrae solo el mensaje limpio

            messages.error(request, error_message)
            return redirect('list_vendedores')

    else:
        # Obtener datos del vendedor actual
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_e_vendedor %s", [idx])
            row = cursor.fetchone()
            if row:
                vendedor = {
                    'ide': row[0],
                    'idvend': row[1],
                    'nombre': row[2],
                    'idalmacen': row[3],
                    'estado': row[5],
                }

    return render(request, 'vendedores_form.html', {'vendedor': vendedor})


@login_required
def delete_vendedor(request, idx):

    with connection.cursor() as cursor:
            cursor.execute("EXEC dj_d_vendedor %s", [idx])

    messages.error(request, 'Eliminado correctamente')
    return redirect('list_vendedores')


# Funcionalidad de rutas

@login_required
def list_ruta(request):
    ide = request.user.ide  # Obtener el 'ide' del usuario autenticado

    with connection.cursor() as cursor:
        cursor.execute("EXEC dj_l_ruta %s", [ide])
        rutas = cursor.fetchall()

    return render(request, 'rutas.html', {'rutas': rutas})

@login_required
def view_ruta(request, idx):
    ide = request.user.ide
    rutas = None

    try:
        vendedores = utils.get_vendedores(ide)
    except Exception as e:
        messages.error(request, f"Error al cargar los vendedores: {str(e)}")

    if request.method == 'POST':
        idvend = request.POST.get('idvend')
        ruta = request.POST.get('ruta')

        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_e_ruta %s, %s, %s, %s",
                               [ide, idvend, ruta, idx])
        messages.success(request, 'Editado correctamente')
        return redirect('list_ruta')  # Redirige después de la edición

    else:
        # Obtener datos del vendedor actual
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_e_ruta %s", [idx])
            row = cursor.fetchone()
            if row:
                rutas = {
                    'idx': row[0],
                    'idvend': row[1],
                    'ide': row[2],
                    'ruta': row[3],
                    'nombre': row[4]
            
                }

    return render(request, 'rutas_view.html', {'rutas': rutas, 'vendedores': vendedores})

@login_required
def add_ruta(request):
    ide = request.user.ide  # Obtener el 'ide' del usuario autenticado
   
    try:
        vendedores = utils.get_vendedores(ide)
    except Exception as e:
        messages.error(request, f"Error al cargar los vendedores: {str(e)}")

    if request.method == 'POST':

        idvend = request.POST.get('idvend')
        ruta = request.POST.get('ruta')

        try:
            with connection.cursor() as cursor:
                    cursor.execute("EXEC dj_a_ruta %s, %s, %s",
                                   [ide, idvend, ruta])

            messages.success(request, 'Agregado correctamente')
            # Redirigir si la inserción es exitosa
            return redirect('list_ruta')

        except Exception as e:
            error_message = str(e)

            # Expresión regular para extraer solo el mensaje de error sin códigos ni detalles técnicos
            match = re.search(r"La ruta.*? ya existe\.", error_message)
            if match:
                error_message = match.group(0)  # Extrae solo el mensaje limpio

            messages.error(request, error_message)

    return render(request, 'rutas_form.html', {'vendedores': vendedores})


@login_required
def edit_ruta(request, idx):
    ide = request.user.ide
    rutas = None

    try:
        vendedores = utils.get_vendedores(ide)
    except Exception as e:
        messages.error(request, f"Error al cargar los vendedores: {str(e)}")

    if request.method == 'POST':
        idvend = request.POST.get('idvend')
        ruta = request.POST.get('ruta')

        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_e_ruta %s, %s, %s, %s",
                               [ide, idvend, ruta, idx])
        messages.success(request, 'Editado correctamente')
        return redirect('list_ruta')  # Redirige después de la edición

    else:
        # Obtener datos del vendedor actual
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_e_ruta %s", [idx])
            row = cursor.fetchone()
            if row:
                rutas = {
                    'idx': row[0],
                    'idvend': row[1],
                    'ide': row[2],
                    'ruta': row[3],
                    'nombre': row[4]
            
                }

    return render(request, 'rutas_form.html', {'rutas': rutas, 'vendedores': vendedores})


@login_required
def delete_ruta(request, idx):

    with connection.cursor() as cursor:
            cursor.execute("EXEC dj_d_ruta %s", [idx])

    messages.error(request, 'Eliminado correctamente')
    return redirect('list_ruta')
    
    
  
@login_required
def reporte_producto_movimientos(request):    
        ide = request.user.ide
        vendedor = []
        results = []
        columns = []
        
        try:
            lineas = utils.get_lineas_producto()
        except Exception as e:
            messages.error(request, f"Error al cargar las lineas: {str(e)}")
 
        try:
            almacen = utils.get_almacen_producto(ide)
        except Exception as e:
            messages.error(request, f"Error al cargar los almacenes: {str(e)}")

        try:
            vendedor = utils.get_vendedores_clientes(ide)
        except Exception as e:
            messages.error(request, f"Error al cargar los vendedores: {str(e)}")

        if "reporte" in request.GET:
            reporte = request.GET.get("reporte")

            # ---- Detallado LISTO ----
            if reporte == "detallado":
                almini = request.GET.get('almini')
                almfin = request.GET.get('almfin')
                codini = request.GET.get('codini')
                codfin = request.GET.get('codfin')
                linini = request.GET.get('linini')
                linfin = request.GET.get('linfin')
                fecha1 = request.GET.get('fecha1')
                fecha2 = request.GET.get('fecha2')
                cliini = request.GET.get('cliini')
                clifin = request.GET.get('clifin')
                proini = request.GET.get('proini')
                profin = request.GET.get('profin')

                with connection.cursor() as cursor:
                    cursor.execute("""
                        EXEC dj_r_deta_movi_prods
                            @ide=%s, 
                            @almini=%s,
                            @almfin=%s, 
                            @codini=%s, 
                            @codfin=%s, 
                            @linini=%s, 
                            @linfin=%s, 
                            @fecha1=%s,
                            @fecha2=%s,
                            @cliini=%s,
                            @clifin=%s, 
                            @proini=%s,
                            @profin=%s
                    """, [
                        ide, almini, almfin, codini, codfin, linini, linfin, fecha1, fecha2, cliini, clifin, proini, profin
                    ])
                    columns = [col[0] for col in cursor.description]
                    results = cursor.fetchall()

            # ------Concentrado Global LISTO-----
            elif reporte == "concentrado_global":
                # Normalizar checkboxes (solo devuelven "on" si están marcados)
                condife = 1 if request.GET.get('condife') == 'on' else 0
                actual = 1 if request.GET.get('actual') == 'on' else 0

                # Variables con nombres correctos
                codprodini = request.GET.get('codini')
                codprofin = request.GET.get('codfin')
                linini = request.GET.get('linini')
                linfin = request.GET.get('linfin')
                fecha1 = request.GET.get('fecha1')
                fecha2 = request.GET.get('fecha2')

                try:
                    with connection.cursor() as cursor:
                        cursor.execute("""
                            EXEC dj_r_concentrado_general
                                @ide=%s,
                                @condife=%s,
                                @actual=%s,
                                @codprodini=%s,
                                @codprodfin=%s,
                                @linini=%s,
                                @linfin=%s,
                                @fecha1=%s,
                                @fecha2=%s
                        """, [ide, condife, actual, codprodini, codprofin, linini, linfin, fecha1, fecha2])

                        columns = [col[0] for col in cursor.description]
                        results = cursor.fetchall()
                  

                except Exception as e:
                    messages.error(request, f"Error al ejecutar el procedimiento: {str(e)}")
                    results = []
                    columns = []

            # --------- Kardex Movimentos LISTO----------------
            elif reporte == "kardex_movimientos":
                
                actual = 1 if request.GET.get('actual') == 'on' else 0
                codigo = request.GET.get('codini')
                fecha_i = request.GET.get('fecha1')
                fecha_f = request.GET.get('fecha2')

                with connection.cursor() as cursor:
                        cursor.execute("""
                            EXEC dj_r_kardex_prod
                                @ide=%s,
                                @actual=%s,
                                @codigo=%s,
                                @fecha_i=%s,
                                @fecha_f=%s
                        """, [ide, actual, codigo, fecha_i, fecha_f])

                        columns = [col[0] for col in cursor.description]
                        results = cursor.fetchall()

            # --------- Reorden total FALTAN LOS DETALLES DE REVISAR ----------------
            elif reporte == "reorden_total":
                
                codini = request.GET.get('codini')
                codfin = request.GET.get('codfin')
                linini = request.GET.get('linini')
                linfin = request.GET.get('linfin')
                cadlin = request.GET.get('cadlin')
                inifact = request.GET.get('inifact')
                inifac2 = request.GET.get('inifac2')
                inifac3 = request.GET.get('inifac3')
                fecha1 = request.GET.get('fecha1')
                fecha2 = request.GET.get('fecha2')

                with connection.cursor() as cursor:
                        cursor.execute("""
                            EXEC dj_r_reordentotal
                                @ide=%s,
                                @codini=%s,
                                @codfin=%s,
                                @linini=%s,
                                @linfin=%s,
                                @cadlin=%s, 
                                @inifact=%s, 
                                @inifac2=%s,
                                @inifac3=%s,
                                @fecha1=%s,
                                @fecha2=%s
                        """, [ide, codini, codfin, linini, linfin, cadlin, inifact, inifac2, inifac3, fecha1, fecha2])

                        columns = [col[0] for col in cursor.description]
                        results = cursor.fetchall()

            # --------- Reorden Almacen PREGUNTAR PORQUE NO APARECEN DATOS Y SOLO CON IDE 3----------------
            elif reporte == "reorden_almacen":
                
                almini = request.GET.get('almini')
                codini = request.GET.get('codini')
                codfin = request.GET.get('codfin')
                linini = request.GET.get('linini')
                linfin = request.GET.get('linfin')
                cadlin = request.GET.get('cadlin')
                inifact = request.GET.get('inifact')
                inifac2 = request.GET.get('inifac2')
                inifac3 = request.GET.get('inifac3')
                fecha1 = parse_fecha(request.GET.get('fecha1'))
                fecha2 = parse_fecha(request.GET.get('fecha2'))

                with connection.cursor() as cursor:
                        cursor.execute("""
                            EXEC dj_r_reordenalmacen
                                @ide=%s,
                                @almini=%s,      
                                @codini=%s,
                                @codfin=%s,
                                @linini=%s,
                                @linfin=%s,
                                @cadlin=%s, 
                                @inifact=%s, 
                                @inifac2=%s,
                                @inifac3=%s,
                                @fecha1=%s,
                                @fecha2=%s
                        """, [ide, almini, codini, codfin, linini, linfin, cadlin, inifact, inifac2, inifac3, fecha1, fecha2])

                        columns = [col[0] for col in cursor.description]
                        results = cursor.fetchall()

             # --------- Compras&Ventas LISTO  ----------------
            elif reporte == "compras&ventas":
                
                idei = request.GET.get('idei')
                idef = request.GET.get('idef')
                linini = request.GET.get('linini')
                linfin = request.GET.get('linfin')
                fecha1 = request.GET.get('fecha1')
                fecha2 = request.GET.get('fecha2')

                with connection.cursor() as cursor:
                        cursor.execute("""
                            EXEC dj_r_compraventa
                                @idei=%s,
                                @idef=%s,      
                                @linini=%s,
                                @linfin=%s,
                                @fecha1=%s,
                                @fecha2=%s
                        """, [idei, idef, linini, linfin, fecha1, fecha2])

                        columns = [col[0] for col in cursor.description]
                        results = cursor.fetchall()

            # --------- Reorden Sucursales NECESITA ARREGLARSE EL PROCEDIMIENTO ALMACENADO----------------
            elif reporte == "reorden_sucursales":
                
                ide_i = request.GET.get('idei')
                ide_f = request.GET.get('idef')
                codini = request.GET.get('codini')
                codfin = request.GET.get('codfin')
                linini = request.GET.get('linini')
                linfin = request.GET.get('linfin')
                cadlin = request.GET.get('cadlin')
                cadide = request.GET.get('cadide')
                inifact = request.GET.get('inifact')
                inifac2 = request.GET.get('inifac2')
                inifac3 = request.GET.get('inifac3')
                fecha1 = request.GET.get('fecha1')
                fecha2 = request.GET.get('fecha2')

                with connection.cursor() as cursor:
                        cursor.execute("""
                            EXEC dj_r_reordentotal_sucursales
                                @ide_i=%s,
                                @ide_f=%s, 
                                @codini=%s,
                                @codfin=%s,     
                                @linini=%s,
                                @linfin=%s,
                                @cadlin=%s,
                                @cadide=%s,
                                @inifact=%s,
                                @inifac2=%s,
                                @inifac3=%s,
                                @fecha1=%s,
                                @fecha2=%s
                        """, [ide_i, ide_f, codini, codfin, linini, linfin, cadlin, cadide, inifact, inifac2, inifac3, fecha1, fecha2])

                        columns = [col[0] for col in cursor.description]
                        results = cursor.fetchall()

                         # --------- Reorden Vendedor ----------------
            elif reporte == "reorden_vendedor":
                
                codini = request.GET.get('codini')
                codfin = request.GET.get('codfin')
                linini = request.GET.get('linini')
                linfin = request.GET.get('linfin')
                vi = request.GET.get('vi')
                vf = request.GET.get('vf')
                cadlin = request.GET.get('cadlin')
                inifact = request.GET.get('inifact')
                inifac2 = request.GET.get('inifac2')
                inifac3 = request.GET.get('inifac3')
                fecha1 = request.GET.get('fecha1')
                fecha2 = request.GET.get('fecha2')

                with connection.cursor() as cursor:
                        cursor.execute("""
                            EXEC dj_r_reordentotal_vend
                                @codini=%s,
                                @codfin=%s,     
                                @linini=%s,
                                @linfin=%s,
                                @vi=%s,
                                @vf=%s,
                                @cadlin=%s,
                                @inifact=%s,
                                @inifac2=%s,
                                @inifac3=%s,
                                @fecha1=%s,
                                @fecha2=%s
                        """, [codini, codfin, linini, linfin, vi, vf, cadlin, inifact, inifac2, inifac3, fecha1, fecha2])

                        columns = [col[0] for col in cursor.description]
                        results = cursor.fetchall()


        return render(request, 'reporte_producto_movimientos.html', {
            'lineas': lineas,
            'almacen': almacen,
            'vendedor': vendedor,
            'columns': columns,
            'results': results
         })

@login_required
def reporte_detallado(request):
    ide = request.user.ide
    results = []
    columns = []
    productos = []

    if "reporte" in request.GET:
        reporte = request.GET.get("reporte")

        if reporte == "detallado":
            almini = request.GET.get('almini')
            almfin = request.GET.get('almfin')
            codini = request.GET.get('codini')
            codfin = request.GET.get('codfin')
            linini = request.GET.get('linini')
            linfin = request.GET.get('linfin')
            fecha1 = request.GET.get('fecha1')
            fecha2 = request.GET.get('fecha2')
            cliini = request.GET.get('cliini')
            clifin = request.GET.get('clifin')
            proini = request.GET.get('proini')
            profin = request.GET.get('profin')
            venini = request.GET.get('venini')
            venfin = request.GET.get('venfin')

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_deta_movi_prods
                        @ide=%s, 
                        @almini=%s,
                        @almfin=%s, 
                        @codini=%s, 
                        @codfin=%s, 
                        @linini=%s, 
                        @linfin=%s, 
                        @fecha1=%s,
                        @fecha2=%s,
                        @cliini=%s,
                        @clifin=%s, 
                        @proini=%s,
                        @profin=%s
                """, [
                    ide, almini, almfin, codini, codfin,
                    linini, linfin, fecha1, fecha2,
                    cliini, clifin, proini, profin
                ])
                columns = [col[0] for col in cursor.description]
                raw_results = cursor.fetchall()
                results = [dict(zip(columns, row)) for row in raw_results]

            # 🔹 Agrupar por producto y movimiento
            productos_dict = defaultdict(lambda: defaultdict(list))
            for r in results:
                productos_dict[r["codigo"]][r["movto"]].append(r)

            productos = []
            for codigo, movtos in productos_dict.items():
                movtos_list = []
                for movto, items in movtos.items():
                    total_cant = sum(Decimal(i["cant"] or 0) for i in items)
                    total_boni = sum(Decimal(i["boni"] or 0) for i in items)
                    movtos_list.append({
                        "movto": movto,
                        "items": items,
                        "total_cant": total_cant,
                        "total_boni": total_boni
                    })
                productos.append({
                    "codigo": codigo,
                    "nombre": next((i["nombre"] for m in movtos.values() for i in m), ""),
                    "movtos": movtos_list
                })

    context = {
        'title': "Reporte Movimientos al Inventario",
        'columns': columns,
        'productos': productos,
        'fecha': datetime.now(),
        'linini': linini if 'linini' in locals() else '',
        'linfin': linfin if 'linfin' in locals() else '',
    }

    pdf = render_to_pdf("reporte_detallado.html", context)
    if pdf:
        return pdf
    return HttpResponse("Error al generar el PDF", status=500)


@login_required
def reporte_concentrado_global(request):
    ide = request.user.ide
    results = []
    columns = []

    if "reporte" in request.GET:
        reporte = request.GET.get("reporte")

        if reporte == "concentrado_global":
                # Normalizar checkboxes (solo devuelven "on" si están marcados)
                condife = 1 if request.GET.get('condife') == 'on' else 0
                actual = 1 if request.GET.get('actual') == 'on' else 0

                # Variables con nombres correctos
                codprodini = request.GET.get('codini')
                codprofin = request.GET.get('codfin')
                linini = request.GET.get('linini')
                linfin = request.GET.get('linfin')
                fecha1 = parse_fecha(request.GET.get('fecha1'))
                fecha2 = parse_fecha(request.GET.get('fecha2'))

                with connection.cursor() as cursor:
                    cursor.execute("""
                        EXEC dj_r_concentrado_general
                            @ide=%s,
                            @condife=%s,
                            @actual=%s,
                            @codprodini=%s,
                            @codprodfin=%s,
                            @linini=%s,
                            @linfin=%s,
                            @fecha1=%s,
                            @fecha2=%s
                        """, [ide, condife, actual, codprodini, codprofin, linini, linfin, fecha1, fecha2])

                    columns = [col[0] for col in cursor.description]
                    raw_results = cursor.fetchall()
                    results = [dict(zip(columns, row)) for row in raw_results]

    context = {
        'title': "Reporte Movimientos al Inventario",
        'columns': columns,
        'results': results,
        'fecha': datetime.now(),
        'linini': linini,
        'linfin': linfin,
        'fecha1': fecha1,
        'fecha2': fecha2
    }

    pdf = render_to_pdf("reporte_concentrado_global.html", context)
    if pdf:
        return pdf
    return HttpResponse("Error al generar el PDF", status=500)


@login_required
def reporte_kardex_movimientos(request):
    ide = request.user.ide
    results = []
    columns = []
    codigo = ''
    fecha_i = ''
    fecha_f = ''

    if "reporte" in request.GET:
        reporte = request.GET.get("reporte")

        if reporte == "kardex_movimientos":
                
                actual = 1 if request.GET.get('actual') == 'on' else 0
                codigo = request.GET.get('codini')
                fecha_i = parse_fecha(request.GET.get('fecha1'))
                fecha_f = parse_fecha(request.GET.get('fecha2'))

                with connection.cursor() as cursor:
                        cursor.execute("""
                            EXEC dj_r_kardex_prod
                                @ide=%s,
                                @actual=%s,
                                @codigo=%s,
                                @fecha_i=%s,
                                @fecha_f=%s
                        """, [ide, actual, codigo, fecha_i, fecha_f])

                        columns = [col[0] for col in cursor.description]
                        raw_results = cursor.fetchall()
                        results = [dict(zip(columns, row)) for row in raw_results]

    context = {
        'title': "Reporte Kardex Movimientos",
        'columns': columns,
        'results': results,
        'codini': codigo,
        'fecha1': fecha_i,
        'fecha2': fecha_f,
        'fecha': datetime.now()
    }

    pdf = render_to_pdf("reporte_kardex_movimientos.html", context)
    if pdf:
        return pdf
    return HttpResponse("Error al generar el PDF", status=500)

@login_required
def reporte_reorden_total(request):
    ide = request.user.ide
    results = []
    columns = []
    fecha1 = ''
    fecha2 = ''

    if "reporte" in request.GET:
        reporte = request.GET.get("reporte")

        if reporte == "reorden_total":
                
                codini = request.GET.get('codini')
                codfin = request.GET.get('codfin')
                linini = request.GET.get('linini')
                linfin = request.GET.get('linfin')
                cadlin = request.GET.get('cadlin')
                inifact = request.GET.get('inifact')
                inifac2 = request.GET.get('inifac2')
                inifac3 = request.GET.get('inifac3')
                fecha1 = parse_fecha(request.GET.get('fecha1'))
                fecha2 = parse_fecha(request.GET.get('fecha2'))

                with connection.cursor() as cursor:
                        cursor.execute("""
                            EXEC dj_r_reordentotal
                                @ide=%s,
                                @codini=%s,
                                @codfin=%s,
                                @linini=%s,
                                @linfin=%s,
                                @cadlin=%s, 
                                @inifact=%s, 
                                @inifac2=%s,
                                @inifac3=%s,
                                @fecha1=%s,
                                @fecha2=%s
                        """, [ide, codini, codfin, linini, linfin, cadlin, inifact, inifac2, inifac3, fecha1, fecha2])

                        columns = [col[0] for col in cursor.description]
                        raw_results = cursor.fetchall()
                        results = [dict(zip(columns, row)) for row in raw_results]

    context = {
        'title': "Reporte para Reorden de Inventario",
        'columns': columns,
        'results': results,
        'linini': linini,
        'linfin': linfin,
        'codini': codini,
        'codfin': codfin,
        'fecha1': fecha1,
        'fecha2': fecha2
    }

    pdf = render_to_pdf("reporte_reorden_total.html", context)
    if pdf:
        return pdf
    return HttpResponse("Error al generar el PDF", status=500)

@login_required
def reporte_reorden_total_excel(request):
                ide = request.user.ide
                codini = request.GET.get('codini')
                codfin = request.GET.get('codfin')
                linini = request.GET.get('linini')
                linfin = request.GET.get('linfin')
                cadlin = request.GET.get('cadlin')
                inifact = request.GET.get('inifact')
                inifac2 = request.GET.get('inifac2')
                inifac3 = request.GET.get('inifac3')
                fecha1 = request.GET.get('fecha1')
                fecha2 = request.GET.get('fecha2')

                with connection.cursor() as cursor:
                        cursor.execute("""
                            EXEC dj_r_reordentotal
                                @ide=%s,
                                @codini=%s,
                                @codfin=%s,
                                @linini=%s,
                                @linfin=%s,
                                @cadlin=%s, 
                                @inifact=%s, 
                                @inifac2=%s,
                                @inifac3=%s,
                                @fecha1=%s,
                                @fecha2=%s
                        """, [ide, codini, codfin, linini, linfin, cadlin, inifact, inifac2, inifac3, fecha1, fecha2])

                        columns = [col[0] for col in cursor.description]
                        results = cursor.fetchall()

                        # Crear libro Excel
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Inventario Reorden Total"

                # Encabezados
                for col_num, column_title in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = column_title
                    cell.font = openpyxl.styles.Font(bold=True)

                # Datos
                for row_num, row_data in enumerate(results, 2):
                    for col_num, cell_value in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col_num).value = cell_value

                # Ajuste de ancho
                for i, col in enumerate(columns, 1):
                    max_length = max(
                        (len(str(cell.value)) for cell in ws[get_column_letter(i)] if cell.value),
                        default=10
                    )
                    ws.column_dimensions[get_column_letter(i)].width = max_length + 2

                # Nombre dinámico del archivo
                nombre_archivo = f"Reporte_Reorden_Total_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

                # Respuesta HTTP
                response = HttpResponse(
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
                wb.save(response)

                return response

@login_required
def revisar_reorden_total(request):
    ide = request.user.ide
    results = []
    columns = []

    codini = request.GET.get('codini')
    codfin = request.GET.get('codfin')
    linini = request.GET.get('linini')
    linfin = request.GET.get('linfin')
    cadlin = request.GET.get('cadlin')
    inifact = request.GET.get('inifact')
    inifac2 = request.GET.get('inifac2')
    inifac3 = request.GET.get('inifac3')
    fecha1 = parse_fecha(request.GET.get('fecha1'))
    fecha2 = parse_fecha(request.GET.get('fecha2'))
                    
    with connection.cursor() as cursor:
                        cursor.execute("""
                            EXEC dj_r_reordentotal
                                @ide=%s,
                                @codini=%s,
                                @codfin=%s,
                                @linini=%s,
                                @linfin=%s,
                                @cadlin=%s, 
                                @inifact=%s, 
                                @inifac2=%s,
                                @inifac3=%s,
                                @fecha1=%s,
                                @fecha2=%s
                        """, [ide, codini, codfin, linini, linfin, cadlin, inifact, inifac2, inifac3, fecha1, fecha2])

                        columns = [col[0] for col in cursor.description]
                        raw_results = cursor.fetchall()
                        results = [dict(zip(columns, row)) for row in raw_results]

    return render(request, 'revisar_orden_total.html', {
        'columns': columns,
        'results': results
        })
   
@login_required 
def reporte_compras_ventas(request):
    results = []
    columns = []
    fecha1 = ''
    fecha2 = ''

    if "reporte" in request.GET:
        reporte = request.GET.get("reporte")

        if reporte == "compras&ventas":
            idei = request.GET.get('idei')
            idef = request.GET.get('idef')
            linini = request.GET.get('linini')
            linfin = request.GET.get('linfin')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_compraventa
                        @idei=%s,
                        @idef=%s,      
                        @linini=%s,
                        @linfin=%s,
                        @fecha1=%s,
                        @fecha2=%s
                """, [idei, idef, linini, linfin, fecha1, fecha2])

                columns = [col[0] for col in cursor.description]
                raw_results = cursor.fetchall()
                results = [dict(zip(columns, row)) for row in raw_results]

            # === 🔢 Calcular totales ===
            total_costototal = sum(r.get('costototal', 0) or 0 for r in results)
            total_ctovtatotal = sum(r.get('ctovtatotal', 0) or 0 for r in results)
            total_ventatotal = sum(r.get('ventatotal', 0) or 0 for r in results)
            total_compratotal = sum(r.get('compratotal', 0) or 0 for r in results)

    context = {
        'title': "Reporte Compras y ventas por Línea",
        'columns': columns,
        'results': results,
        'linini': linini,
        'linfin': linfin,
        'fecha1': fecha1,
        'fecha2': fecha2,
        # 👇 pasamos los totales al contexto
        'total_costototal': total_costototal if results else 0,
        'total_ctovtatotal': total_ctovtatotal if results else 0,
        'total_ventatotal': total_ventatotal if results else 0,
        'total_compratotal': total_compratotal if results else 0,
    }

    pdf = render_to_pdf("reporte_compras_ventas.html", context)
    if pdf:
        return pdf
    return HttpResponse("Error al generar el PDF", status=500)

@login_required
def reporte_reorden_vendedor(request):
    results = []
    columns = []
    fecha1 = ''
    fecha2 = ''

    if "reporte" in request.GET:
        reporte = request.GET.get("reporte")

        if reporte == "reorden_vendedor":
                
                codini = request.GET.get('codini')
                codfin = request.GET.get('codfin')
                linini = request.GET.get('linini')
                linfin = request.GET.get('linfin')
                vi = request.GET.get('vi')
                vf = request.GET.get('vf')
                cadlin = request.GET.get('cadlin')
                inifact = request.GET.get('inifact')
                inifac2 = request.GET.get('inifac2')
                inifac3 = request.GET.get('inifac3')
                fecha1 = request.GET.get('fecha1')
                fecha2 = request.GET.get('fecha2')

                with connection.cursor() as cursor:
                        cursor.execute("""
                            EXEC dj_r_reordentotal_vend
                                @codini=%s,
                                @codfin=%s,     
                                @linini=%s,
                                @linfin=%s,
                                @vi=%s,
                                @vf=%s,
                                @cadlin=%s,
                                @inifact=%s,
                                @inifac2=%s,
                                @inifac3=%s,
                                @fecha1=%s,
                                @fecha2=%s
                        """, [codini, codfin, linini, linfin, vi, vf, cadlin, inifact, inifac2, inifac3, fecha1, fecha2])

                        columns = [col[0] for col in cursor.description]
                        raw_results = cursor.fetchall()
                        results = [dict(zip(columns, row)) for row in raw_results]

    context = {
        'title': "Reporte para Reorden de Inventario",
        'columns': columns,
        'results': results,
        'linini': linini,
        'linfin': linfin,
        'codini': codini,
        'codfin': codfin,
        'fecha1': fecha1,
        'fecha2': fecha2
    }

    pdf = render_to_pdf("reporte_reorden_total.html", context)
    if pdf:
        return pdf
    return HttpResponse("Error al generar el PDF", status=500)


# Funcionalidad de grupos

@login_required
def list_grupos(request):
    with connection.cursor() as cursor:
        # Ejecutar el procedimiento correcto
        cursor.execute("EXEC dj_l_grupo %s", ['LIST'])
        grupos = cursor.fetchall()
           
    if not grupos:
        grupos = [] 

    return render(request, 'grupos.html', {'grupos': grupos})

@login_required
def view_grupos(request, idx):
    """
    Visualizar el grupo existente.
    """
    if request.method == 'POST':
        idcp = request.POST.get('idcp')
        nombre = request.POST.get('nombre')
        grupo = request.POST.get('grupo')
        with connection.cursor() as cursor:
                    cursor.execute("EXEC dj_e_grupo %s, %s, %s, %s", [
                                   idcp, nombre, grupo, idx])

        messages.success(request, 'Editado correctamente')
        return redirect('list_grupos')

    else:
        # Obtener datos del vendedor actual
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_e_grupo %s", [idx])
            row = cursor.fetchone()
            if row:
                grupos = {
                    'idcp': row[1],
                    'nombre': row[2],
                    'grupo': row[3]
                }

    return render(request, 'grupos_view.html', {'grupos': grupos})

@login_required
def add_grupos(request):
    """
    Agrega un nuevo grupo.
    """
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        grupo = request.POST.get('grupo')

        with connection.cursor() as cursor:
                    cursor.execute("EXEC dj_a_grupo %s, %s",
                                   [nombre, grupo])

        messages.success(request, 'Agregado correctamente')
        # Redirigir si la inserción es exitosa
        return redirect('list_grupos')

    return render(request, 'grupos_form.html')


@login_required
def edit_grupos(request, idx):
    """
    Edita el grupo existente.
    """
    if request.method == 'POST':
        idcp = request.POST.get('idcp')
        nombre = request.POST.get('nombre')
        grupo = request.POST.get('grupo')
        with connection.cursor() as cursor:
                    cursor.execute("EXEC dj_e_grupo %s, %s, %s, %s", [
                                   idcp, nombre, grupo, idx])

        messages.success(request, 'Editado correctamente')
        return redirect('list_grupos')

    else:
        # Obtener datos del vendedor actual
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_e_grupo %s", [idx])
            row = cursor.fetchone()
            if row:
                grupos = {
                    'idcp': row[1],
                    'nombre': row[2],
                    'grupo': row[3]
                }

    return render(request, 'grupos_form.html', {'grupos': grupos})


@login_required
def delete_grupos(request, idx):
    """
    Elimina un grupo.
    """
    with connection.cursor() as cursor:
        # Llamada al procedimiento almacenado
        cursor.execute("EXEC dj_d_grupo @Idx = %s", [idx])

    messages.error(request, 'Eliminado correctamente')
    return redirect('list_grupos')

# Funcionalidad de lineas


@login_required
def list_lineas(request):
   # Ejecutar el procedimiento almacenado
    with connection.cursor() as cursor:
        cursor.execute("EXEC dj_l_lineas")
        linea = cursor.fetchall()  # Obtener los resultados de la consulta
        
    return render(request, 'lineas.html', {'lineas': linea})

@login_required
def view_lineas(request, idx):
    grupos = []  # Lista para almacenar los grupos
    lineas = []

    try:
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_a_lineas %s", ['list'])
            grupos = cursor.fetchall()  # Devuelve una lista de tuplas (idcp, nombre)
    except Exception as e:
        messages.error(request, f"Error al cargar grupos: {str(e)}")

    if request.method == 'POST':
        idcp = request.POST.get('idcp')
        if not idcp:
            messages.error(request, "Debe seleccionar un grupo.")
            return render(request, "lineas.html")
        
        idcl = request.POST.get('idcl')
        descripcion = request.POST.get('descripcion')
        linea = request.POST.get('linea')
        descto = float(request.POST.get('descto', 0))
        boni = float(request.POST.get('boni', 0))
        f1 = int(request.POST.get('f1', 0))
        f2 = int(request.POST.get('f2', 0))
        f3 = int(request.POST.get('f3', 0))
        
        with connection.cursor() as cursor:
                    cursor.execute(
                        "EXEC dj_e_lineas %s, %s, %s, %s, %s, %s, %s, %s, %s, %s",
                        [idcp, idcl, descripcion, linea, descto, boni, f1, f2, f3, idx]
                    )
                    
                    messages.success(request, "Línea editada correctamente")
                    return redirect("list_lineas")
    else:
        # Obtener datos del vendedor actual
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_e_lineas %s", [idx])
            row = cursor.fetchone()
            if row:
                lineas = {
                    'idcp': row[1],
                    'idcl': row[2],
                    'descripcion': row[3],
                    'linea': row[4],
                    'descto': row[5],
                    'boni': row[6],
                    'f1': row[7],
                    'f2': row[8],
                    'f3': row[9]
                }
                
    return render(request, "lineas_view.html", {'grupos':grupos, 'lineas':lineas})


@login_required
def add_lineas(request):
   
    grupos = []  # Lista para almacenar los grupos

    try:
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_a_lineas %s", ['list'])
            grupos = cursor.fetchall()  # Devuelve una lista de tuplas (idcp, nombre)
    except Exception as e:
        messages.error(request, f"Error al cargar grupos: {str(e)}")

    if request.method == 'POST':
        idcp = request.POST.get('idcp')
        if not idcp:
            messages.error(request, "Debe seleccionar un grupo.")
            return render(request, "lineas.html")
        
        descripcion = request.POST.get('descripcion')
        linea = request.POST.get('linea')
        f1 = int(request.POST.get('f1', 0))
        f2 = int(request.POST.get('f2', 0))
        f3 = int(request.POST.get('f3', 0))
        
        with connection.cursor() as cursor:
                    cursor.execute(
                        "EXEC dj_a_lineas %s, %s, %s, %s, %s, %s",
                        [idcp, descripcion, linea, f1, f2, f3]
                    )
                    
                    messages.success(request, "Línea agregada correctamente")
                    return redirect("list_lineas")

    return render(request, "lineas_form.html", {'grupos':grupos})


@login_required
def edit_lineas(request, idx):
    grupos = []  # Lista para almacenar los grupos
    lineas = []

    try:
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_a_lineas %s", ['list'])
            grupos = cursor.fetchall()  # Devuelve una lista de tuplas (idcp, nombre)
    except Exception as e:
        messages.error(request, f"Error al cargar grupos: {str(e)}")

    if request.method == 'POST':
        idcp = request.POST.get('idcp')
        if not idcp:
            messages.error(request, "Debe seleccionar un grupo.")
            return render(request, "lineas.html")
        descripcion = request.POST.get('descripcion')
        linea = request.POST.get('linea')
        f1 = int(request.POST.get('f1', 0))
        f2 = int(request.POST.get('f2', 0))
        f3 = int(request.POST.get('f3', 0))
        
        with connection.cursor() as cursor:
                    cursor.execute(
                        "EXEC dj_e_lineas %s, %s, %s, %s, %s, %s, %s",
                        [idcp, descripcion, linea, f1, f2, f3, idx]
                    )
                    
                    messages.success(request, "Línea editada correctamente")
                    return redirect("list_lineas")
    else:
        # Obtener datos del vendedor actual
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_e_lineas %s", [idx])
            row = cursor.fetchone()
            if row:
                lineas = {
                    'idcp': row[1],
                    'idcl': row[2],
                    'descripcion': row[3],
                    'linea': row[4],
                    'f1': row[5],
                    'f2': row[6],
                    'f3': row[7]
                }
                
    return render(request, "lineas_form.html", {'grupos':grupos, 'lineas':lineas})

@login_required
def delete_lineas(request, idx):
    """
    Elimina una línea.
    """

    with connection.cursor() as cursor:
        # Llamada al procedimiento almacenado
        cursor.execute("EXEC dj_d_lineas @Idx = %s", [idx])
            # Confirmar si se eliminó correctamente
        return redirect('list_lineas')  # Redirigir después de la eliminación


# Funcionalidad de choferes

@login_required
def list_choferes(request):
    ide = request.user.ide  # Obtener el 'ide' del usuario autenticado

    with connection.cursor() as cursor:
        cursor.execute("EXEC dj_l_chofer %s", [ide])
        choferes = cursor.fetchall()

    return render(request, 'choferes.html', {'choferes': choferes})

@login_required
def view_choferes(request, idx):
    ide = request.user.ide
    choferes = None

    if request.method == 'POST':

        idc = request.POST.get('idc')
        chofer = request.POST.get('chofer')
        licencia = request.POST.get('licencia')
        tipo = request.POST.get('tipo')
        estado = request.POST.get('estado')
        rfc = request.POST.get('rfc')
        domicilio = request.POST.get('domicilio')
        ac = request.POST.get('ac')
        vence = request.POST.get('vence')
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_e_chofer %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s", [
                               ide, idc, chofer, licencia, tipo, estado, rfc, domicilio, ac, vence, idx])

        messages.success(request, 'Editado correctamente')
        return redirect('list_choferes')  # Redirige después de la edición

    else:
        # Obtener datos del vendedor actual
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_e_chofer %s", [idx])
            row = cursor.fetchone()
            if row:
                choferes = {
                    'ide': row[0],
                    'idc': row[1],
                    'chofer': row[2],
                    'licencia': row[3],
                    'tipo': row[4],
                    'estado': row[5],
                    'rfc': row[7],
                    'domicilio': row[8],
                    'ac': row[9],
                    'vence': row[10]
                }
    return render(request, 'choferes_view.html', {'chofer': choferes})


@login_required
def add_choferes(request):
    ide = request.user.ide
    """
    Agrega un nuevo choferes.
    """
    if request.method == 'POST':
        idc = request.POST.get('idc')
        chofer = request.POST.get('chofer')
        licencia = request.POST.get('licencia')
        tipo = request.POST.get('tipo')
        estado = request.POST.get('estado')
        rfc = request.POST.get('rfc')
        domicilio = request.POST.get('domicilio')
        ac = request.POST.get('ac')
        vence = request.POST.get('vence')

        try:
            with connection.cursor() as cursor:
                    cursor.execute("EXEC dj_a_chofer %s, %s, %s, %s, %s, %s, %s, %s, %s, %s",
               [ide, idc, chofer, licencia, tipo, estado, rfc, domicilio, vence, ac])

            messages.success(request, 'Agregado correctamente')
            # Redirigir si la inserción es exitosa
            return redirect('list_choferes')

        except Exception as e:
            error_message = str(e)

            # Expresión regular para extraer solo el mensaje de error sin códigos ni detalles técnicos
            match = re.search(r"El chofer.*? ya existe\.", error_message)
            if match:
                error_message = match.group(0)  # Extrae solo el mensaje limpio

            messages.error(request, error_message)
    return render(request, 'choferes_form.html')


@login_required
def edit_choferes(request, idx):
    ide = request.user.ide
    choferes = None

    if request.method == 'POST':

        idc = request.POST.get('idc')
        chofer = request.POST.get('chofer')
        licencia = request.POST.get('licencia')
        tipo = request.POST.get('tipo')
        estado = request.POST.get('estado')
        rfc = request.POST.get('rfc')
        domicilio = request.POST.get('domicilio')
        ac = request.POST.get('ac')
        vence = request.POST.get('vence')
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_e_chofer %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s", [
                               ide, idc, chofer, licencia, tipo, estado, rfc, domicilio, ac, vence, idx])

        messages.success(request, 'Editado correctamente')
        return redirect('list_choferes')  # Redirige después de la edición

    else:
        # Obtener datos del vendedor actual
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_e_chofer %s", [idx])
            row = cursor.fetchone()
            if row:
                choferes = {
                    'ide': row[0],
                    'idc': row[1],
                    'chofer': row[2],
                    'licencia': row[3],
                    'tipo': row[4],
                    'estado': row[5],
                    'rfc': row[7],
                    'domicilio': row[8],
                    'ac': row[9],
                    'vence': row[10]
                }
    return render(request, 'choferes_form.html', {'chofer': choferes})


@login_required
def delete_choferes(request, idx):

    with connection.cursor() as cursor:
            cursor.execute("EXEC dj_d_chofer %s", [idx])

    messages.error(request, 'Eliminado correctamente')
    return redirect('list_choferes')


# Funcionalidad de tipos de entradas / salidas

@login_required
def list_ensas(request):
    """
    Lista de todos las entradas y salidas.
    """
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_l_tipoensa %s', ['LIST'])
        ensa = cursor.fetchall()

    return render(request, 'ensas.html', {'ensas': ensa})

@login_required
def view_ensas(request, idx):
    ensa = None

    if request.method == 'POST':
        idtes = request.POST.get('idtes')
        nombre = request.POST.get('nombre')
        s = request.POST.get('s')
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_e_tipoensa %s, %s, %s, %s",
                           [idtes, nombre, s, idx])

        messages.success(request, "Editado correctamente")
        return redirect('list_ensas')

    else:
        # Obtener datos del vendedor actual
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_e_tipoensa %s", [idx])
            row = cursor.fetchone()
            if row:
                ensa = {
                    'idtes': row[0],
                    'nombre': row[1],
                    's': row[3]
                }

    return render(request, 'ensas_view.html', {'ensa': ensa})


@login_required
def add_ensas(request):
    """
    Agrega una nueva entrada/salida.
    """
    if request.method == 'POST':
        idtes = request.POST.get('idtes')
        nombre = request.POST.get('nombre')
        s = request.POST.get('s')

        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_a_tipoensa %s, %s, %s',
                        [idtes, nombre, s])

            # Mensaje de éxito
        messages.success(request, "Nueva Entrada/Salida agregada con éxito.")
        return redirect('list_ensas')

    return render(request, 'ensas_form.html')


@login_required
def edit_ensas(request, idx):

    ensa = None

    if request.method == 'POST':
        idtes = request.POST.get('idtes')
        nombre = request.POST.get('nombre')
        s = request.POST.get('s')
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_e_tipoensa %s, %s, %s, %s",
                           [idtes, nombre, s, idx])

        messages.success(request, "Editado correctamente")
        return redirect('list_ensas')

    else:
        # Obtener datos del vendedor actual
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_e_tipoensa %s", [idx])
            row = cursor.fetchone()
            if row:
                ensa = {
                    'idtes': row[0],
                    'nombre': row[1],
                    's': row[3]
                }

    return render(request, 'ensas_form.html', {'ensa': ensa})


@login_required
def delete_ensas(request, idx):
    """
    Elimina la entrada y salida existente.
    """
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_d_tipoensa %s', [idx])

    messages.error(request, "Eliminado correctamente")
    return redirect('list_ensas')

# Funcionalidad de mensajes para clientes


@login_required
def list_mens_clientes(request):
    """
    Lista de todos los mensajes de los clientes.
    """
    ide = request.user.ide  # Obtener el 'ide' del usuario autenticado

    with connection.cursor() as cursor:
        cursor.execute("EXEC dj_l_mens_cliente %s", [ide])
        mens_clientes = cursor.fetchall()

    return render(request, 'mens_clientes.html', {'mens_clientes': mens_clientes})

@login_required
def view_mens_clientes(request, idx):
    ide = request.user.ide
    mens_cliente = None

    if request.method == 'POST':
        idmc = request.POST.get('idmc')
        nombre = request.POST.get('nombre')

        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_e_mens_cliente %s, %s, %s, %s", [
                           idmc, nombre, ide, idx])

        messages.success(request, 'Editado correctamente')
        return redirect('list_mens_clientes')

    else:
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_e_mens_cliente %s", [idx])
            row = cursor.fetchone()

            if row:
                mens_cliente = {
                    # Asegúrate de que los índices son correctos
                    'idmc': row[0],
                    'nombre': row[1],
                    'ide': row[3]
                }

    return render(request, 'mens_clientes_view.html', {'mens_cliente': mens_cliente})

@login_required
def add_mens_clientes(request):
    ide = request.user.ide  # Obtener el 'ide' del usuario autenticado
    """
    Agrega un nuevo mensaje para el cliente.
    """
    if request.method == 'POST':
        nombre = request.POST.get('nombre')

        try:
            with connection.cursor() as cursor:
                    cursor.execute("EXEC dj_a_mens_cliente %s, %s", [nombre, ide])

            messages.success(request, 'Agregado correctamente')
            return redirect('list_mens_clientes')

        except Exception as e:
            error_message = str(e)

            # Expresión regular para extraer solo el mensaje de error sin códigos ni detalles técnicos
            match = re.search(r"El mensaje .*? ya existe\.", error_message)
            if match:
                error_message = match.group(0)  # Extrae solo el mensaje limpio

            messages.error(request, error_message)

    return render(request, 'mens_clientes_form.html')


@login_required
def edit_mens_clientes(request, idx):
    ide = request.user.ide
    mens_cliente = None

    if request.method == 'POST':
        idmc = request.POST.get('idmc')
        nombre = request.POST.get('nombre')

        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_e_mens_cliente %s, %s, %s, %s", [
                           idmc, nombre, ide, idx])

        messages.success(request, 'Editado correctamente')
        return redirect('list_mens_clientes')

    else:
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_e_mens_cliente %s", [idx])
            row = cursor.fetchone()

            if row:
                mens_cliente = {
                    # Asegúrate de que los índices son correctos
                    'idmc': row[0],
                    'nombre': row[1],
                    'ide': row[3]
                }

    return render(request, 'mens_clientes_form.html', {'mens_cliente': mens_cliente})


@login_required
def delete_mens_clientes(request, idx):

    with connection.cursor() as cursor:
            cursor.execute("EXEC dj_d_mens_cliente %s", [idx])
    messages.error(request, 'Eliminado correctamente')
    return redirect('list_mens_clientes')


# Funcionalidad de transportes

@login_required
def list_transps(request):
    ide = request.user.ide
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_l_transp %s', [ide])
        transps = cursor.fetchall()

    return render(request, 'transps.html', {'transps': transps})

@login_required
def view_transps(request, idx):
    ide = request.user.ide
   
    if request.method == 'POST':
        idv = request.POST.get('idv')
        transp = request.POST.get('transp')
        color = request.POST.get('color')
        ano = request.POST.get('ano')
        placas = request.POST.get('placas')
        serie = request.POST.get('serie')
        asegura = request.POST.get('asegura')
        poliza = request.POST.get('poliza')
        inciso = request.POST.get('inciso')
        vence = request.POST.get('vence')
        cve_sat = request.POST.get('cve_sat')
        peso = request.POST.get('peso')
        ac = request.POST.get('ac')
        tc = request.POST.get('tc')
        emision = request.POST.get('emision')
        lugar = request.POST.get('lugar')
        tipo = request.POST.get('tipo')

        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_e_transp %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s ,%s ,%s, %s, %s", [
                                   ide, idv, transp, color, ano, placas, serie, asegura, poliza, inciso, vence, cve_sat, peso, ac, tc, emision, lugar, tipo, idx])

            messages.success(request, 'Editado correctamente')
            return redirect('list_transps')
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_transp %s", [idx])
                row = cursor.fetchone()
        if row:
                transp = {
                    # Asegúrate de que los índices son correctos
                    'idv': row[1],
                    'transp': row[2],
                    'color': row[3],
                    'ano': row[4],
                    'placas': row[5],
                    'serie': row[6],
                    'asegura': row[7],
                    'poliza': row[8],
                    'inciso': row[9],
                    'vence': row[10],
                    'cve_sat': row[12],
                    'peso': row[13],
                    'ac': row[14],
                    'tc': row[15],
                    'emision': row[16],
                    'lugar': row[17],
                    'tipo': row[18]          
                }
        
    return render(request, 'transps_view.html', {'transp': transp})

@login_required
def add_transps(request):
    ide = request.user.ide
    if request.method == 'POST':
            idv = request.POST.get('idv')
            transp = request.POST.get('transp')
            color = request.POST.get('color')
            ano = request.POST.get('ano')
            placas = request.POST.get('placas')
            serie = request.POST.get('serie')
            asegura = request.POST.get('asegura')
            poliza = request.POST.get('poliza')
            inciso = request.POST.get('inciso')
            vence = request.POST.get('vence')
            cve_sat = request.POST.get('cve_sat')
            peso = request.POST.get('peso')
            ac = request.POST.get('ac')
     
            try:
                with connection.cursor() as cursor:
                    cursor.execute("EXEC dj_a_transp %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s", [
                                   ide, idv, transp, color, ano, placas, serie, asegura, poliza, inciso, vence, cve_sat, peso, ac])

                messages.success(request, 'Agregado correctamente')
                return redirect('list_transps')

            except Exception as e:
             error_message = str(e)

                # Expresión regular para extraer solo el mensaje de error sin códigos ni detalles técnicos
             match = re.search(r"El vehiculo con código .*? ya existe\.", error_message)
            if match:
                    # Extrae solo el mensaje limpio
                    error_message = match.group(0)

            messages.error(request, error_message)

    return render(request, 'transps_form.html')

@login_required
def edit_transps(request, idx):
    ide = request.user.ide
   
    if request.method == 'POST':
        idv = request.POST.get('idv')
        transp = request.POST.get('transp')
        color = request.POST.get('color')
        ano = request.POST.get('ano')
        placas = request.POST.get('placas')
        serie = request.POST.get('serie')
        asegura = request.POST.get('asegura')
        poliza = request.POST.get('poliza')
        inciso = request.POST.get('inciso')
        vence = request.POST.get('vence')
        cve_sat = request.POST.get('cve_sat')
        peso = request.POST.get('peso')
        ac = request.POST.get('ac')

        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_e_transp %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s", [
                                   ide, idv, transp, color, ano, placas, serie, asegura, poliza, inciso, vence, cve_sat, peso, ac, idx])

            messages.success(request, 'Editado correctamente')
            return redirect('list_transps')
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_transp %s", [idx])
                row = cursor.fetchone()
        if row:
                transp = {
                    # Asegúrate de que los índices son correctos
                    'idv': row[1],
                    'transp': row[2],
                    'color': row[3],
                    'ano': row[4],
                    'placas': row[5],
                    'serie': row[6],
                    'asegura': row[7],
                    'poliza': row[8],
                    'inciso': row[9],
                    'vence': row[10],
                    'cve_sat': row[12],
                    'peso': row[13],
                    'ac': row[14]        
                }
        
    return render(request, 'transps_form.html', {'transp': transp})

@login_required
def delete_transps(request, idx):
    """
    Elimina los transportes existentes.
    """
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_d_transp %s', [idx])
    
    messages.error(request, 'Eliminado correctamente')
    return redirect('list_transps')


# Funcionalidad de calendario de reparto

@login_required
def search_repartos(request):
    ide = request.user.ide
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_l_reparto %s', [ide])
        repartos = cursor.fetchall()

    return render(request, 'repartos.html', {'repartos': repartos})

@login_required
def add_repartos(request):
    ide = request.user.ide
    if request.method == 'POST':
            idv = request.POST.get('idv')
            transp = request.POST.get('transp')
            color = request.POST.get('color')
            ano = request.POST.get('ano')
            placas = request.POST.get('placas')
            serie = request.POST.get('serie')
            asegura = request.POST.get('asegura')
            poliza = request.POST.get('poliza')
            inciso = request.POST.get('inciso')
            vence = request.POST.get('vence')
            cve_sat = request.POST.get('cve_sat')
            peso = request.POST.get('peso')
            ac = request.POST.get('ac')
            tc = request.POST.get('tc')
            emision = request.POST.get('emision')
            lugar = request.POST.get('lugar')
            tipo = request.POST.get('tipo')
            
            try:
                with connection.cursor() as cursor:
                    cursor.execute("EXEC dj_a_transp %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s ,%s ,%s, %s", [
                                   ide, idv, transp, color, ano, placas, serie, asegura, poliza, inciso, vence, cve_sat, peso, ac, tc, emision, lugar, tipo])

                messages.success(request, 'Agregado correctamente')
                return redirect('list_repartos')

            except Exception as e:
             error_message = str(e)

                # Expresión regular para extraer solo el mensaje de error sin códigos ni detalles técnicos
             match = re.search(r"El vehiculo con código .*? ya existe\.", error_message)
            if match:
                    # Extrae solo el mensaje limpio
                    error_message = match.group(0)

            messages.error(request, error_message)
    return render(request, 'repartos_form.html')


@login_required
def edit_repartos(request, idx):
    ide = request.user.ide
    if request.method == 'POST':
        idv = request.POST.get('idv')
        transp = request.POST.get('transp')
        color = request.POST.get('color')
        ano = request.POST.get('ano')
        placas = request.POST.get('placas')
        serie = request.POST.get('serie')
        asegura = request.POST.get('asegura')
        poliza = request.POST.get('poliza')
        inciso = request.POST.get('inciso')
        vence = request.POST.get('vence')
        cve_sat = request.POST.get('cve_sat')
        peso = request.POST.get('peso')
        ac = request.POST.get('ac')
        tc = request.POST.get('tc')
        emision = request.POST.get('emision')
        lugar = request.POST.get('lugar')
        tipo = request.POST.get('tipo')

        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_e_transp %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s ,%s ,%s, %s, %s", [
                                   ide, idv, transp, color, ano, placas, serie, asegura, poliza, inciso, vence, cve_sat, peso, ac, tc, emision, lugar, tipo, idx])

            messages.success(request, 'Editado correctamente')
            return redirect('list_repartos')
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_transp %s", [idx])
                row = cursor.fetchone()
        if row:
                reparto = {
                    # Asegúrate de que los índices son correctos
                    'idv': row[1],
                    'transp': row[2],
                    'color': row[3],
                    'ano': row[4],
                    'placas': row[5],
                    'serie': row[6],
                    'asegura': row[7],
                    'poliza': row[8],
                    'inciso': row[9],
                    'vence': row[10],
                    'cve_sat': row[12],
                    'peso': row[13],
                    'ac': row[14],
                    'tc': row[15],
                    'emision': row[16],
                    'lugar': row[17],
                    'tipo': row[18]          
                }
    return render(request, 'repartos_form.html', {'reparto': reparto})


@login_required
def delete_repartos(request, idx):
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_d_transp %s', [idx])
    
    messages.error(request, 'Eliminado correctamente')
    return redirect('list_repartos')

# Funcionalidad de tipos de clientes

@login_required
def list_tipoclies(request):
    
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_l_tipoclie %s', ['LIST'])
        tipoclies = cursor.fetchall()
        
    return render(request, 'tipoclies.html', {'tipoclies': tipoclies})

@login_required
def view_tipoclies(request, idx):
    """
    Editar tipos de clientes existente.
    """
    if request.method == 'POST':
        idtc = request.POST.get('idtc')
        nombre = request.POST.get('nombre')
        
        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_e_tipoclie %s, %s, %s', [idtc, nombre, idx])
        
        messages.success(request, 'Editado correctamente')
        return redirect('list_tipoclies')
    
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_tipoclie %s", [idx])
                row = cursor.fetchone()
        if row:
                tipoclie = {
                    # Asegúrate de que los índices son correctos
                    'idtc': row[0],
                    'nombre': row[1]
                }
    
    return render(request, 'tipoclies_view.html', {'tipoclie': tipoclie})

@login_required
def add_tipoclies(request):
    
    if request.method == 'POST':
            nombre=request.POST.get('nombre')
    try:       
        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_a_tipoclie %s', [nombre])
        
        messages.success(request, 'Agregado correctamente')
        return redirect('list_tipoclies')

    except Exception as e:
        error_message = str(e)

        # Expresión regular para extraer solo el mensaje de error sin códigos ni detalles técnicos
        match = re.search(r"El tipo de cliente con código .*? ya existe\.", error_message)
    if match:
        # Extrae solo el mensaje limpio
        error_message = match.group(0)

        messages.error(request, error_message)
    return render(request, 'tipoclies_form.html')


@login_required
def edit_tipoclies(request, idx):
    """
    Editar tipos de clientes existente.
    """
    if request.method == 'POST':
        idtc = request.POST.get('idtc')
        nombre = request.POST.get('nombre')
        
        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_e_tipoclie %s, %s, %s', [idtc, nombre, idx])
        
        messages.success(request, 'Editado correctamente')
        return redirect('list_tipoclies')
    
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_tipoclie %s", [idx])
                row = cursor.fetchone()
        if row:
                tipoclie = {
                    # Asegúrate de que los índices son correctos
                    'idtc': row[0],
                    'nombre': row[1]
                }
    
    return render(request, 'tipoclies_form.html', {'tipoclie': tipoclie})


@login_required
def delete_tipoclies(request, idx):
    """
    Elimina los tipos de clientes existentes.
    """
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_d_tipoclie %s', [idx])
        
    messages.error(request, 'Eliminado correctamente')
    return redirect('list_tipoclies')


# Funcionamiento de embarques para clientes

@login_required
def list_embarques(request):
    ide = request.user.ide
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_l_embarque %s', [ide])
        embarques = cursor.fetchall()
    
    return render(request, 'embarques.html', {'embarques': embarques})

@login_required
def view_embarques(request, idx):
    ide = request.user.ide
    if request.method == 'POST':
        idem = request.POST.get('idem')  
        nombre = request.POST.get('nombre') 

        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_e_embarque %s, %s, %s, %s', [ide, idem, nombre, idx])
        
        messages.success(request, 'Editado correctamente')
        return redirect('list_embarques')
    
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_embarque %s", [idx])
                row = cursor.fetchone()
        if row:
                embarque = {
                    # Asegúrate de que los índices son correctos
                    'idem': row[2],
                    'nombre': row[3]
                }
    return render(request, 'embarques_view.html', {'embarque': embarque})

@login_required
def add_embarques(request):
    ide = request.user.ide

    if request.method == 'POST':
        nombre = request.POST.get('nombre')

        try:
            with connection.cursor() as cursor:
                cursor.execute('EXEC dj_a_embarque %s, %s', [ide, nombre])

            messages.success(request, 'Agregado correctamente')
            return redirect('list_embarques')

        except Exception as e:
            error_message = str(e)
            match = re.search(r"El tipo de embarque con código .*? ya existe\.|No hay códigos disponibles.*", error_message)
            if match:
                error_message = match.group(0)

            messages.error(request, error_message)

    return render(request, 'embarques_form.html')
  

@login_required
def edit_embarques(request, idx):
    ide = request.user.ide
    if request.method == 'POST':
        idem = request.POST.get('idem')  
        nombre = request.POST.get('nombre') 

        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_e_embarque %s, %s, %s, %s', [ide, idem, nombre, idx])
        
        messages.success(request, 'Editado correctamente')
        return redirect('list_embarques')
    
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_embarque %s", [idx])
                row = cursor.fetchone()
        if row:
                embarque = {
                    # Asegúrate de que los índices son correctos
                    'idem': row[2],
                    'nombre': row[3]
                }
    return render(request, 'embarques_form.html', {'embarque': embarque})


@login_required
def delete_embarques(request, idx):
   
    with connection.cursor() as cursor:
       cursor.execute('EXEC dj_d_embarque %s', [idx])
    
    messages.error(request, 'Eliminado correctamente')
    return redirect('list_embarques')


# Funcionalidad de cuentas por pagar

@login_required
def list_aplicacxps(request):
    ide = request.user.ide
    
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_l_aplicacxp %s', [ide])
        aplicacxps = cursor.fetchall()
        
    return render(request, 'aplicacxp.html', {'aplicacxps': aplicacxps})

@login_required
def view_aplicacxps(request, idx):
    ide = request.user.ide
    if request.method == 'POST':
        idaplica = request.POST.get('idaplica')  
        nombre = request.POST.get('nombre')  # <- Sin la coma
        
        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_e_aplicacxp %s, %s, %s, %s', [ide, idaplica, nombre, idx])
        
        messages.success(request, 'Editado correctamente')
        return redirect('list_aplicacxps')
    
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_aplicacxp %s", [idx])
                row = cursor.fetchone()
        if row:
                aplicacxp = {
                    # Asegúrate de que los índices son correctos
                    'idaplica': row[2],
                    'nombre': row[3]
                }
    return render(request, 'aplicacxp_view.html', {'aplicacxp': aplicacxp})

@login_required
def add_aplicacxps(request):
        ide = request.user.ide
        if request.method == 'POST':
            nombre=request.POST.get('nombre')
        try:
            with connection.cursor() as cursor:
                cursor.execute('EXEC dj_a_aplicacxp %s, %s', [ide, nombre])   
                
            messages.success(request, 'Agregado correctamente')     
            return redirect('list_aplicacxps')
        
        except Exception as e:
            error_message = str(e)

            match = re.search(r"El tipo de cuenta con código .*? ya existe\.", error_message)
        if match:
            error_message = match.group(0)
            messages.error(request, error_message)
            
        return render(request, 'aplicacxp_form.html')


@login_required
def edit_aplicacxps(request, idx):
    ide = request.user.ide
    if request.method == 'POST':
        idaplica = request.POST.get('idaplica')  
        nombre = request.POST.get('nombre')  # <- Sin la coma
        
        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_e_aplicacxp %s, %s, %s, %s', [ide, idaplica, nombre, idx])
        
        messages.success(request, 'Editado correctamente')
        return redirect('list_aplicacxps')
    
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_aplicacxp %s", [idx])
                row = cursor.fetchone()
        if row:
                aplicacxp = {
                    # Asegúrate de que los índices son correctos
                    'idaplica': row[2],
                    'nombre': row[3]
                }
    return render(request, 'aplicacxp_form.html', {'aplicacxp': aplicacxp})


@login_required
def delete_aplicacxps(request, idx):
    
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_d_aplicacxp %s', [idx])
        
    messages.error(request, 'Eliminado correctamente')
    return redirect('list_aplicacxps')


# Funcionalidad de tipos de cargos y devoluciones para proveedores

@login_required
def list_carabocxps(request):
    ide = request.user.ide
    
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_l_carabocxp %s', [ide])
        carabocxps = cursor.fetchall()
    
    return render(request, 'carabocxp.html', {'carabocxps': carabocxps})

@login_required
def view_carabocxps(request, idx):
    ide = request.user.ide
    if request.method == 'POST':
        idcacxp = request.POST.get('idcacxp')  # <- Sin la coma
        nombre = request.POST.get('nombre')  # <- Sin la coma

        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_e_carabocxp %s, %s, %s, %s', [ide, idcacxp, nombre, idx])
        
        messages.success(request, 'Editado correctamente')
        return redirect('list_carabocxps')
    
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_carabocxp %s", [idx])
                row = cursor.fetchone()
        if row:
                carabocxp = {
                    # Asegúrate de que los índices son correctos
                    'idcacxp': row[2],
                    'nombre': row[3]
                }
    return render(request, 'carabocxp_view.html', {'carabocxp': carabocxp})

@login_required
def add_carabocxps(request):
        ide = request.user.ide
        if  request.method == 'POST':
            idcacxp=request.POST.get('idcacxp')
            nombre=request.POST.get('nombre')
        try:
            with connection.cursor() as cursor:
                cursor.execute('EXEC dj_a_carabocxp %s, %s, %s', [ide, idcacxp, nombre])
                
            messages.success(request,'Agregado correctamente')
            return redirect('list_carabocxps')
        
        except Exception as e:
            error_message = str(e)

            match = re.search(r"El tipo de cuenta con código .*? ya existe\.", error_message)
        if  match:
            error_message = match.group(0)
            messages.error(request, error_message)
        
            return redirect('list_carabocxps')
        return render(request, 'carabocxp_form.html')


@login_required
def edit_carabocxps(request, idx):
    ide = request.user.ide
    if request.method == 'POST':
        idcacxp = request.POST.get('idcacxp')  # <- Sin la coma
        nombre = request.POST.get('nombre')  # <- Sin la coma

        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_e_carabocxp %s, %s, %s, %s', [ide, idcacxp, nombre, idx])
        
        messages.success(request, 'Editado correctamente')
        return redirect('list_carabocxps')
    
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_carabocxp %s", [idx])
                row = cursor.fetchone()
        if row:
                carabocxp = {
                    # Asegúrate de que los índices son correctos
                    'idcacxp': row[2],
                    'nombre': row[3]
                }
    return render(request, 'carabocxp_form.html', {'carabocxp': carabocxp})


@login_required
def delete_carabocxps(request, idx):
    
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_d_carabocxp %s', [idx])
    
    messages.error(request, 'Eliminado correctamente')
    return redirect('list_carabocxps')


# Funcionalidad de proveedores

@login_required
def list_proveedor(request):
    ide = request.user.ide
    
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_l_proveedo %s', [ide])
        proveedor = cursor.fetchall()
    
    return render(request, 'proveedor.html', {'proveedores': proveedor})

@login_required
def view_proveedor(request, idx):
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_a_proveedor %s", ['list'])
            proveedor1 = cursor.fetchall()
    except Exception as e:
        messages.error(request, f"Error al cargar grupos: {str(e)}")
        proveedor1 = []
    
    ide = request.user.ide
    if request.method == 'POST':
        # Aquí deberías capturar todos los campos definidos en tu SP
        idcp = request.POST.get('idcp', '')
        nombre = request.POST.get('nombre', '')
        razon = request.POST.get('razon', '')
        rfc = request.POST.get('rfc', '')
        domicilio = request.POST.get('domicilio', '')
        colonia = request.POST.get('colonia', '')
        cp = request.POST.get('cp', '')
        poblacion = request.POST.get('poblacion', '')
        email = request.POST.get('email', '')
        telefono = request.POST.get('telefono', '')
        fax = request.POST.get('fax', '')
        credito = request.POST.get('credito') or 0
        dg = request.POST.get('dg') or 0
        dp = request.POST.get('dp') or 0
        idpt = request.POST.get('idpt','')
        lab = request.POST.get('lab') == '1'
        cxp = request.POST.get('cxp') == '1'
        ventas = request.POST.get('ventas', '')
        telventas = request.POST.get('telventas', '')
        c_vtas = request.POST.get('c_vtas', '')
        cobranza = request.POST.get('cobranza', '')
        telcobranza = request.POST.get('telcobranza', '')
        c_cob = request.POST.get('c_cob', '')
        otro = request.POST.get('otro', '')
        telotro = request.POST.get('telotro', '')
        c_otro = request.POST.get('c_otro', '')
        otro1 = request.POST.get('otro1', '')
        telotro1 = request.POST.get('telotro1', '')
        c_otro1 = request.POST.get('c_otro1', '')
        b1 = request.POST.get('b1', '')
        c1 = request.POST.get('c1', '')
        cb1 = request.POST.get('cb1', '')
        s1 = request.POST.get('s1', '')
        r1 = request.POST.get('r1', '')
        b2 = request.POST.get('b2', '')
        c2 = request.POST.get('c2', '')
        cb2 = request.POST.get('cb2', '')
        s2 = request.POST.get('s2', '')
        r2 = request.POST.get('r2', '')
        b3 = request.POST.get('b3', '')
        c3 = request.POST.get('c3', '')
        cb3 = request.POST.get('cb3', '')
        s3 = request.POST.get('s3', '')
        r3 = request.POST.get('r3', '')
        b4 = request.POST.get('b4', '')
        c4 = request.POST.get('c4', '')
        cb4 = request.POST.get('cb4', '')
        s4 = request.POST.get('s4', '')
        r4 = request.POST.get('r4', '')
        
        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_e_proveedo %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s',
                    [ide, idx, idcp, nombre, razon, rfc, domicilio, colonia, cp, poblacion, email, telefono, fax, credito, dg, dp, idpt, lab, cxp, ventas, telventas, c_vtas, cobranza, telcobranza, c_cob, otro, telotro, c_otro, otro1, telotro1, c_otro1,
                     b1, c1, cb1, s1, r1, b2, c2, cb2, s2, r2, b3, c3, cb3, s3, r3, b4, c4, cb4, s4, r4])
        
        messages.success(request, 'Editado correctamente')
        return redirect('list_proveedor')
    
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_proveedo %s", [idx])
                row = cursor.fetchone()
        if row:
                proveedor = {
                    # Asegúrate de que los índices son correctos
                    'nombre': row[3],
                    'idcp': row[2],
                    'razon': row[4],
                    'rfc': row[9],
                    'domicilio': row[5],
                    'colonia': row[6],
                    'cp': row[7],
                    'poblacion': row[8],
                    'email': row[10],
                    'telefono': row[11],
                    'fax': row[12],
                    'credito': row[13],
                    'dg': row[14],
                    'dp': row[15],
                    'idpt': row[16],
                    'lab': row[17],
                    'cxp': row[18],
                    'ventas': row[19],
                    'telventas': row[20],
                    'c_vtas': row[21],
                    'cobranza': row[22],
                    'telcobranza': row[23],
                    'c_cob': row[24],
                    'otro': row[25],
                    'telotro': row[26],
                    'c_otro': row[27],
                    'otro1': row[28],
                    'telotro1': row[29],
                    'c_otro1': row[30],
                    'b1': row[31],
                    'c1': row[32],
                    'cb1': row[33],
                    's1': row[34],
                    'r1': row[35],
                    'b2': row[36],
                    'c2': row[37],
                    'cb2': row[38],
                    's2': row[39],
                    'r2': row[40],
                    'b3': row[41],
                    'c3': row[42],
                    'cb3': row[43],
                    's3': row[44],
                    'r3': row[45],
                    'b4': row[46],
                    'c4': row[47],
                    'cb4': row[48],
                    's4': row[49],
                    'r4': row[50],
                    'notas': row[51],
                }
    return render(request, 'proveedor_view.html', {'proveedor': proveedor, 'proveedor1': proveedor1})

@login_required
def add_proveedor(request):
    try:
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_a_proveedor %s", ['list'])
            proveedor1 = cursor.fetchall()
    except Exception as e:
        messages.error(request, f"Error al cargar grupos: {str(e)}")
        proveedor1 = []

    ide = request.user.ide

    if request.method == 'POST':
        # Obtener datos del formulario
          # Supongamos que recibes los parámetros a través de un formulario o API
        nombre = request.POST.get('nombre', '')
        razon = request.POST.get('razon', '')
        rfc = request.POST.get('rfc', '')
        domicilio = request.POST.get('domicilio', '')
        colonia = request.POST.get('colonia', '')
        cp = request.POST.get('cp', '')
        poblacion = request.POST.get('poblacion', '')
        email = request.POST.get('email', '')
        telefono = request.POST.get('telefono', '')
        fax = request.POST.get('fax', '')
        try:
            credito = int(request.POST.get('credito', '0') or 0)
        except ValueError:
            credito = 0  # O maneja como quieras si es inválido
        dg = request.POST.get('dg', '0.0')  # Asegúrate de que sea un número (float)
        dp = request.POST.get('dp', '0.0')  # Asegúrate de que sea un número (float)
        idpt = request.POST.get('idpt', '')
        lab = request.POST.get('lab', 'False')  # Convertir a booleano (True/False)
        cxp = request.POST.get('cxp', 'False')  # Convertir a booleano (True/False)
        ventas = request.POST.get('ventas', '')
        telventas = request.POST.get('telventas', '')
        c_vtas = request.POST.get('c_vtas', '')
        cobranza = request.POST.get('cobranza', '')
        telcobranza = request.POST.get('telcobranza', '')
        c_cob = request.POST.get('c_cob', '')
        otro = request.POST.get('otro', '')
        telotro = request.POST.get('telotro', '')
        c_otro = request.POST.get('c_otro', '')
        otro1 = request.POST.get('otro1', '')
        telotro1 = request.POST.get('telotro1', '')
        c_otro1 = request.POST.get('c_otro1', '')
        b1 = request.POST.get('b1', '')
        c1 = request.POST.get('c1', '')
        cb1 = request.POST.get('cb1', '')
        s1 = request.POST.get('s1', '')
        r1 = request.POST.get('r1', '')
        b2 = request.POST.get('b2', '')
        c2 = request.POST.get('c2', '')
        cb2 = request.POST.get('cb2', '')
        s2 = request.POST.get('s2', '')
        r2 = request.POST.get('r2', '')
        b3 = request.POST.get('b3', '')
        c3 = request.POST.get('c3', '')
        cb3 = request.POST.get('cb3', '')
        s3 = request.POST.get('s3', '')
        r3 = request.POST.get('r3', '')
        b4 = request.POST.get('b4', '')
        c4 = request.POST.get('c4', '')
        cb4 = request.POST.get('cb4', '')
        s4 = request.POST.get('s4', '')
        r4 = request.POST.get('r4', '')

        try:
            with connection.cursor() as cursor:
                placeholders = ', '.join(['%s'] * 49)  # 49 parámetros
                sql = f'EXEC dj_a_proveedo {placeholders}'
                cursor.execute(sql, [
                    nombre, razon, rfc, domicilio, colonia, cp, poblacion,
                    email, telefono, fax, credito, dg, dp, idpt, lab, cxp,
                    ventas, telventas, c_vtas, 
                    cobranza, telcobranza, c_cob,
                    otro, telotro, c_otro, 
                    otro1, telotro1, c_otro1,
                    b1, c1, cb1, s1, r1,
                    b2, c2, cb2, s2, r2,
                    b3, c3, cb3, s3, r3,
                    b4, c4, cb4, s4, r4, ide
                ])
            messages.success(request, 'Proveedor agregado correctamente.')
            return redirect('list_proveedor')

        except Exception as e:
            error_message = str(e)
            match = re.search(r"El proveedor con código .*? ya existe\.", error_message)
            if match:
                messages.error(request, match.group(0))
            else:
                messages.error(request, 'Error al agregar proveedor: ' + error_message)
            return redirect('add_proveedor')

    return render(request, 'proveedor_form.html', {'proveedor1': proveedor1})

@login_required
def edit_proveedor(request, idx):
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_a_proveedor %s", ['list'])
            proveedor1 = cursor.fetchall()
    except Exception as e:
        messages.error(request, f"Error al cargar grupos: {str(e)}")
        proveedor1 = []
    
    ide = request.user.ide
    if request.method == 'POST':
        # Aquí deberías capturar todos los campos definidos en tu SP
        idcp = request.POST.get('idcp', '')
        nombre = request.POST.get('nombre', '')
        razon = request.POST.get('razon', '')
        rfc = request.POST.get('rfc', '')
        domicilio = request.POST.get('domicilio', '')
        colonia = request.POST.get('colonia', '')
        cp = request.POST.get('cp', '')
        poblacion = request.POST.get('poblacion', '')
        email = request.POST.get('email', '')
        telefono = request.POST.get('telefono', '')
        fax = request.POST.get('fax', '')
        credito = request.POST.get('credito') or 0
        dg = request.POST.get('dg') or 0
        dp = request.POST.get('dp') or 0
        idpt = request.POST.get('idpt','')
        lab = request.POST.get('lab') == '1'
        cxp = request.POST.get('cxp') == '1'
        ventas = request.POST.get('ventas', '')
        telventas = request.POST.get('telventas', '')
        c_vtas = request.POST.get('c_vtas', '')
        cobranza = request.POST.get('cobranza', '')
        telcobranza = request.POST.get('telcobranza', '')
        c_cob = request.POST.get('c_cob', '')
        otro = request.POST.get('otro', '')
        telotro = request.POST.get('telotro', '')
        c_otro = request.POST.get('c_otro', '')
        otro1 = request.POST.get('otro1', '')
        telotro1 = request.POST.get('telotro1', '')
        c_otro1 = request.POST.get('c_otro1', '')
        b1 = request.POST.get('b1', '')
        c1 = request.POST.get('c1', '')
        cb1 = request.POST.get('cb1', '')
        s1 = request.POST.get('s1', '')
        r1 = request.POST.get('r1', '')
        b2 = request.POST.get('b2', '')
        c2 = request.POST.get('c2', '')
        cb2 = request.POST.get('cb2', '')
        s2 = request.POST.get('s2', '')
        r2 = request.POST.get('r2', '')
        b3 = request.POST.get('b3', '')
        c3 = request.POST.get('c3', '')
        cb3 = request.POST.get('cb3', '')
        s3 = request.POST.get('s3', '')
        r3 = request.POST.get('r3', '')
        b4 = request.POST.get('b4', '')
        c4 = request.POST.get('c4', '')
        cb4 = request.POST.get('cb4', '')
        s4 = request.POST.get('s4', '')
        r4 = request.POST.get('r4', '')
        
        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_e_proveedo %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s',
                    [ide, idx, idcp, nombre, razon, rfc, domicilio, colonia, cp, poblacion, email, telefono, fax, credito, dg, dp, idpt, lab, cxp, ventas, telventas, c_vtas, cobranza, telcobranza, c_cob, otro, telotro, c_otro, otro1, telotro1, c_otro1,
                     b1, c1, cb1, s1, r1, b2, c2, cb2, s2, r2, b3, c3, cb3, s3, r3, b4, c4, cb4, s4, r4])
        
        messages.success(request, 'Editado correctamente')
        return redirect('list_proveedor')
    
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_proveedo %s", [idx])
                row = cursor.fetchone()
        if row:
                proveedor = {
                    # Asegúrate de que los índices son correctos
                    'nombre': row[3],
                    'idcp': row[2],
                    'razon': row[4],
                    'rfc': row[9],
                    'domicilio': row[5],
                    'colonia': row[6],
                    'cp': row[7],
                    'poblacion': row[8],
                    'email': row[10],
                    'telefono': row[11],
                    'fax': row[12],
                    'credito': row[13],
                    'dg': row[14],
                    'dp': row[15],
                    'idpt': row[16],
                    'lab': row[17],
                    'cxp': row[18],
                    'ventas': row[19],
                    'telventas': row[20],
                    'c_vtas': row[21],
                    'cobranza': row[22],
                    'telcobranza': row[23],
                    'c_cob': row[24],
                    'otro': row[25],
                    'telotro': row[26],
                    'c_otro': row[27],
                    'otro1': row[28],
                    'telotro1': row[29],
                    'c_otro1': row[30],
                    'b1': row[31],
                    'c1': row[32],
                    'cb1': row[33],
                    's1': row[34],
                    'r1': row[35],
                    'b2': row[36],
                    'c2': row[37],
                    'cb2': row[38],
                    's2': row[39],
                    'r2': row[40],
                    'b3': row[41],
                    'c3': row[42],
                    'cb3': row[43],
                    's3': row[44],
                    'r3': row[45],
                    'b4': row[46],
                    'c4': row[47],
                    'cb4': row[48],
                    's4': row[49],
                    'r4': row[50],
                    'notas': row[51],
                }
    return render(request, 'proveedor_form.html', {'proveedor': proveedor, 'proveedor1': proveedor1})


@login_required
def delete_proveedor(request, idx):
    
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_d_proveedo %s', [idx])
    
    messages.error(request, 'Eliminado correctamente')
    return redirect('list_proveedor')

# Funcionalidad de ubicaciones

@login_required
def list_ubicacion(request):
    ide = request.user.ide
    
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_l_ubicacion %s', [ide])
        ubicacion = cursor.fetchall()
    
    return render(request, 'ubicacion.html', {'ubicaciones': ubicacion})

@login_required
def add_ubicacion(request):
    ide = request.user.ide

    if request.method == 'POST':
        idubica = request.POST.get('idubica')
        print('idubica:', idubica)

        if not idubica:
            messages.error(request, 'El campo Ubicación no puede estar vacío.')
            return render(request, 'ubicacion_form.html', {'ubicaciones': None})

        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_a_ubicacion %s, %s', [ide, idubica])

        messages.success(request, 'Agregado correctamente')
        return redirect('list_ubicacion')

    return render(request, 'ubicacion_form.html', {'ubicaciones': None})

@login_required
def view_ubicacion(request, idx):
   
    if request.method == 'POST':
        idubica = request.POST.get('idubica')  # <- Sin la coma

        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_e_ubicacion %s, %s', [ idubica, idx])
        
        messages.success(request, 'Editado correctamente')
        return redirect('list_ubicacion')
    
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_ubicacion %s", [idx])
                row = cursor.fetchone()
        if row:
                ubicacion = {
                    # Asegúrate de que los índices son correctos
                    'idubica': row[2]
                }
    return render(request, 'ubicacion_view.html', {'ubicaciones': ubicacion})



@login_required
def edit_ubicacion(request, idx):
   
    if request.method == 'POST':
        idubica = request.POST.get('idubica')  # <- Sin la coma

        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_e_ubicacion %s, %s', [ idubica, idx])
        
        messages.success(request, 'Editado correctamente')
        return redirect('list_ubicacion')
    
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_ubicacion %s", [idx])
                row = cursor.fetchone()
        if row:
                ubicacion = {
                    # Asegúrate de que los índices son correctos
                    'idubica': row[2]
                }
    return render(request, 'ubicacion_form.html', {'ubicaciones': ubicacion})


@login_required
def delete_ubicacion(request, idx):
    
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_d_ubicacion %s', [idx])
        
        messages.error(request, 'Eliminado correctamente')
    return redirect('list_ubicacion')
        

# Funcionalidad de clientes

@login_required 
def list_cliente(request):
    ide = request.user.ide  # Obtener el 'ide' del usuario autenticado
    cliini = request.GET.get('cliini', '10')
    clifin = request.GET.get('clifin', '20')
    cliente = []

    if cliini and clifin:
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_cliente %s, %s, %s", [ide, cliini, clifin])
            cliente = cursor.fetchall()

    return render(request, 'cliente.html', {'clientes': cliente, 'cliini': cliini, 'clifin': clifin})

def parse_fecha(fecha_str):
    if not fecha_str:
        return None
    try:
        if "/" in fecha_str:
            return datetime.strptime(fecha_str, "%d/%m/%Y")  # Devuelve datetime
        elif "-" in fecha_str:
            return datetime.strptime(fecha_str, "%Y-%m-%d")
        elif len(fecha_str) == 8:
            return datetime.strptime(fecha_str, "%Y%m%d")
        else:
            return None
    except Exception:
        return None

    
@login_required
def reporte_cliente(request):
    ide = request.user.ide
    results = []
    columns = []
    lineas = []

    try:
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_cliente_linea")
            lineas = cursor.fetchall() 
    except Exception as e:
        messages.error(request, f"Error al cargar las lineas: {str(e)}")

    try:
        with connection.cursor() as cursor:
            cursor.execute("EXEC dj_l_vendedor %s", [ide])
            vendedor = cursor.fetchall()
    except Exception as e:
        messages.error(request, f"Error al cargar los vendedores: {str(e)}")

    if "reporte" in request.GET:
        reporte = request.GET.get("reporte")

        # ---- Antigüedad de saldos ----
        if reporte == "antiguedad":
            cini = request.GET.get('cini')
            cfin = request.GET.get('cfin')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))
            rini = request.GET.get('rini')
            rfin = request.GET.get('rfin')
            fini = request.GET.get('fini')
            ffin = request.GET.get('ffin')
            fechapago = parse_fecha(request.GET.get('fechapago'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_ctes_antiguedasaldos 
                        @ide=%s, @cini=%s, @cfin=%s, 
                        @fecha1=%s, @fecha2=%s, 
                        @rini=%s, @rfin=%s, 
                        @fini=%s, @ffin=%s, 
                        @fechapago=%s
                """, [
                    ide, cini, cfin,
                    fecha1, fecha2,
                    rini, rfin,
                    fini, ffin,
                    fechapago
                ])
                columns = [col[0] for col in cursor.description]
                results = cursor.fetchall()

        elif reporte == "estado_detallado":
            cini = request.GET.get('cini')
            cfin = request.GET.get('cfin')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))
            rini = request.GET.get('rini')
            rfin = request.GET.get('rfin')

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_ctes_edoctadetallado 
                        @ide=%s, @cini=%s, @cfin=%s, 
                        @fecha1=%s, @fecha2=%s, 
                        @rini=%s, @rfin=%s
                """, [
                    ide, cini, cfin,
                    fecha1, fecha2,
                    rini, rfin
                ])
                columns = [col[0] for col in cursor.description]
                results = cursor.fetchall()

        # ---- Estado de Cuenta Detallado ----
        elif reporte == "catalogo_clientes":
            cini = request.GET.get('cini')
            cfin = request.GET.get('cfin')
            rini = request.GET.get('rini')
            rfin = request.GET.get('rfin')

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_clientes 
                        @ide=%s, @cini=%s, @cfin=%s, 
                        @rini=%s, @rfin=%s
                """, [
                    ide, cini, cfin, rini, rfin
                ])
                columns = [col[0] for col in cursor.description]
                results = cursor.fetchall()
        
        # ---- Antiguedad Saldos Contabilidad ----

        elif reporte == "antiguedad_conta":
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))
            fechapago = parse_fecha(request.GET.get('fechapago'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_antiguedadsaldoscontabilidad 
                        @fecha1=%s, @fecha2=%s, @fechapago=%s
                """, [
                    fecha1, fecha2, fechapago
                ])
                columns = [col[0] for col in cursor.description]
                results = cursor.fetchall()
        
        # ---- Estado de Cuenta Gral ----
         
        elif reporte == "estado_general":
            cini = request.GET.get('cini')
            cfin = request.GET.get('cfin')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))
            rini = request.GET.get('rini')
            rfin = request.GET.get('rfin')
            fini = request.GET.get('fini')
            ffin = request.GET.get('ffin')

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_ctes_edoctagral
                        @ide=%s, @cini=%s, @cfin=%s, 
                        @fecha1=%s, @fecha2=%s, 
                        @rini=%s, @rfin=%s, 
                        @fini=%s, @ffin=%s
                """, [
                    ide, cini, cfin,
                    fecha1, fecha2,
                    rini, rfin,
                    fini, ffin
                ])
                columns = [col[0] for col in cursor.description]
                results = cursor.fetchall()

         # ---- Recuperacion de cartera ----

        elif reporte == "recuperacion_cartera":
            cini = request.GET.get('cini')
            cfin = request.GET.get('cfin')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))
            rini = request.GET.get('rini')
            rfin = request.GET.get('rfin')
            fini = request.GET.get('fini')
            ffin = request.GET.get('ffin')
            fechapago = parse_fecha(request.GET.get('fechapago'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_ctes_recuperacion
                        @ide=%s, @cini=%s, @cfin=%s, 
                        @fecha1=%s, @fecha2=%s, 
                        @rini=%s, @rfin=%s, 
                        @fini=%s, @ffin=%s, 
                        @fechapago=%s
                """, [
                    ide, cini, cfin,
                    fecha1, fecha2,
                    rini, rfin,
                    fini, ffin,
                    fechapago
                ])
                columns = [col[0] for col in cursor.description]
                results = cursor.fetchall()

             # ---- Antiguedad Saldos Linea Alimentos ----
             
        elif reporte == "linea_alimentos":
            cini = request.GET.get('cini')
            cfin = request.GET.get('cfin')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))
            rini = request.GET.get('rini')
            rfin = request.GET.get('rfin')
            fini = request.GET.get('fini')
            ffin = request.GET.get('ffin')
            idclin = request.GET.get('idclin')
            fechapago = parse_fecha(request.GET.get('fechapago'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_ctes_antiguedasaldos_alimentos
                        @ide=%s, @cini=%s, @cfin=%s, 
                        @fecha1=%s, @fecha2=%s, 
                        @rini=%s, @rfin=%s, 
                        @fini=%s, @ffin=%s, 
                        @idclin=%s,@fechapago=%s
                """, [
                    ide, cini, cfin,
                    fecha1, fecha2,
                    rini, rfin,
                    fini, ffin, idclin, fechapago
                ])
                columns = [col[0] for col in cursor.description]
                results = cursor.fetchall()

                print(columns)
                print(results)

        # ---- Resumen de ventas, Devs y cuotas de Vendedores ----
             
        elif reporte == "ventas":
            venini = request.GET.get('venini')
            venfin = request.GET.get('venfin')
            linini = request.GET.get('linini')
            linfin = request.GET.get('linfin' , '99999')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_gralvtas_dev_ven
                        @ide=%s, @venini=%s, @venfin=%s, @linini=%s, @linfin=%s, 
                        @fecha1=%s, @fecha2=%s
                """, [
                    ide, venini, venfin, linini, linfin,
                    fecha1, fecha2
                ])
                columns = [col[0] for col in cursor.description]
                results = cursor.fetchall()

        # ----- Resumen de ventas, devs y Utilidades de Vendedores -----

        elif reporte == "resumen":
            venini = request.GET.get('venini')
            venfin = request.GET.get('venfin')
            linini = request.GET.get('linini')
            linfin = request.GET.get('linfin' , '99999')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_gralvtas_utils_dev_ven
                        @ide=%s, @venini=%s, @venfin=%s, @linini=%s, @linfin=%s, 
                        @fecha1=%s, @fecha2=%s
                """, [
                    ide, venini, venfin, linini, linfin,
                    fecha1, fecha2
                ])
                columns = [col[0] for col in cursor.description]
                results = cursor.fetchall()
        
        # ---- Saldos de clientes 80/20 ----

        elif reporte == "8020":
            grp8020 = request.GET.get('grp8020')
            cini = request.GET.get('cini')
            cfin = request.GET.get('cfin')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))
            rini = request.GET.get('rini')
            rfin = request.GET.get('rfin')
            fini = request.GET.get('fini')
            ffin = request.GET.get('ffin')
            fechapago = parse_fecha(request.GET.get('fechapago'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_info_vtas_80_20_cte 
                        @grp8020=%s, @ide=%s, @cini=%s, @cfin=%s, 
                        @fecha1=%s, @fecha2=%s, 
                        @rini=%s, @rfin=%s, 
                        @fini=%s, @ffin=%s, 
                        @fechapago=%s
                """, [
                    grp8020, ide, cini, cfin,
                    fecha1, fecha2,
                    rini, rfin,
                    fini, ffin,
                    fechapago
                ])
                columns = [col[0] for col in cursor.description]
                results = cursor.fetchall()

    return render(request, "reporte_cliente.html", {
        "columns": columns,
        "results": results,
        'lineas': lineas,
        'vendedor': vendedor
    })

@login_required
def reporte_antiguedad_cliente_pdf(request):
    ide = request.user.ide
    results = []
    columns = []

    if "reporte" in request.GET:
        reporte = request.GET.get("reporte")

        if reporte == "antiguedad":
            cini = request.GET.get('cini')
            cfin = request.GET.get('cfin')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))
            rini = request.GET.get('rini')
            rfin = request.GET.get('rfin')
            fini = request.GET.get('fini')
            ffin = request.GET.get('ffin')
            fechapago = parse_fecha(request.GET.get('fechapago'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_ctes_antiguedasaldos 
                        @ide=%s, @cini=%s, @cfin=%s, 
                        @fecha1=%s, @fecha2=%s, 
                        @rini=%s, @rfin=%s, 
                        @fini=%s, @ffin=%s, 
                        @fechapago=%s
                """, [
                    ide, cini, cfin,
                    fecha1, fecha2,
                    rini, rfin,
                    fini, ffin,
                    fechapago
                ])
                columns = [col[0] for col in cursor.description]
                raw_results = cursor.fetchall()

            # Convertir cada row en dict
            results = [dict(zip(columns, row)) for row in raw_results]

            # Filtrar: solo filas con algún saldo
            results = [
                r for r in results
                if any(r.get(col, 0) for col in columns[8:16])  # asumo col 9 a 14 = saldos
            ]

            # Calcular subtotales por cliente
            subtotales = defaultdict(lambda: {col: 0 for col in columns[8:16]})
            for r in results:
                cliente = r[columns[2]]  # row.2 = cliente
                for col in columns[8:16]:
                    subtotales[cliente][col] += r.get(col, 0)

            # Construir lista enriquecida con subtotales
            enriched = []
            for i, r in enumerate(results):
                enriched.append(r)

                # Verificar si es la última fila de este cliente
                next_cliente = results[i + 1][columns[2]] if i + 1 < len(results) else None
                if r[columns[2]] != next_cliente:
                    enriched.append({
                        "subtotal": True,
                        "cliente": r[columns[2]],
                        **subtotales[r[columns[2]]]
                    })
            results = enriched

    context = {
        "title": "Reporte de Antiguedad de Saldos de Clientes",
        "fecha": datetime.now(),
        "columns": columns,
        "results": results,
        "cini": cini,
        "cfin": cfin,
        "fecha1": fecha1,
        "fecha2": fecha2, 
        "fechapago": fechapago
    }

    pdf = render_to_pdf("reporte_antiguedad_cliente_pdf.html", context)
    if pdf:
        return pdf
    return HttpResponse("Error al generar el PDF", status=500)


def reporte_estado_detallado_pdf(request):
    ide = request.user.ide

    results1, results2, columns1, columns2, totales = [], [], [], [], {}

    if "reporte" in request.GET:
        reporte = request.GET.get("reporte")

        if reporte == "estado_detallado":
            cini = request.GET.get("cini")
            cfin = request.GET.get("cfin")
            fecha1 = parse_fecha(request.GET.get("fecha1"))
            fecha2 = parse_fecha(request.GET.get("fecha2"))
            rini = request.GET.get("rini")
            rfin = request.GET.get("rfin")

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_ctes_edoctadetallado  
                        @ide=%s, @cini=%s, @cfin=%s, 
                        @fecha1=%s, @fecha2=%s, 
                        @rini=%s, @rfin=%s
                """, [
                    ide, cini, cfin,
                    fecha1, fecha2,
                    rini, rfin
                ])

                # --- Primer resultset: clientes
                columns1 = [col[0] for col in cursor.description]
                results1 = [dict(zip(columns1, row)) for row in cursor.fetchall()]

                # --- Segundo resultset: detalle ventas/notas
                if cursor.nextset():
                    columns2 = [col[0] for col in cursor.description]
                    results2 = [dict(zip(columns2, row)) for row in cursor.fetchall()]

            # --- Calcular totales por cliente
            totales = defaultdict(lambda: {"cargo": 0, "abono": 0, "saldo": 0})
            for row in results2:
                cid = row["idcliente"]
                totales[cid]["cargo"] += row.get("cargos", 0) or 0
                totales[cid]["abono"] += row.get("abonos", 0) or 0
                totales[cid]["saldo"] += row.get("saldo", 0) or 0

    context = {
        "title": "Estado de Cuenta Detallado",
        "fecha": datetime.now(),
        "results1": results1,   # clientes
        "results2": results2,   # detalle de movimientos
        "columns1": columns1,
        "columns2": columns2,
        "cini": cini,
        "cfin": cfin,
        "fecha1": fecha1,
        "fecha2": fecha2,
        "totales": totales,
    }

    pdf = render_to_pdf("reporte_estado_detallado_pdf.html", context)
    if pdf:
        return pdf
    return HttpResponse("Error al generar el PDF", status=500)

def reporte_antiguedad_conta_excel(request):
    fecha1 = request.GET.get('fecha1')
    fecha2 = request.GET.get('fecha2')
    fechapago = request.GET.get('fechapago')

    # Ejecutamos el procedimiento almacenado
    with connection.cursor() as cursor:
        cursor.execute("""
            EXEC dj_r_antiguedadsaldoscontabilidad 
                @fecha1=%s, @fecha2=%s, @fechapago=%s
        """, [fecha1, fecha2, fechapago])
        columns = [col[0] for col in cursor.description]
        results = cursor.fetchall()

    # Crear libro Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Antigüedad Saldos Contabilidad"

    # Escribir encabezados
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = openpyxl.styles.Font(bold=True)

    # Escribir datos
    for row_num, row_data in enumerate(results, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = cell_value

    # Ajustar ancho de columnas
    for i, col in enumerate(columns, 1):
        max_length = max(
            (len(str(cell.value)) for cell in ws[get_column_letter(i)] if cell.value),
            default=10
        )
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    # Nombre dinámico del archivo
    nombre_archivo = f"Antiguedad_Saldos_Contabilidad_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Preparar respuesta HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)

    return response


@login_required
def reporte_linea_alimentos_pdf(request):
    ide = request.user.ide
    results, columns = [], []

    if "reporte" in request.GET:
        reporte = request.GET.get("reporte")

        if reporte == "linea_alimentos":
            cini = request.GET.get('cini')
            cfin = request.GET.get('cfin')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))
            rini = request.GET.get('rini')
            rfin = request.GET.get('rfin')
            fini = request.GET.get('fini')
            ffin = request.GET.get('ffin')
            idclin = request.GET.get('idclin')
            fechapago = parse_fecha(request.GET.get('fechapago'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_ctes_antiguedasaldos_alimentos  
                        @ide=%s, @cini=%s, @cfin=%s, 
                        @fecha1=%s, @fecha2=%s, 
                        @rini=%s, @rfin=%s, 
                        @fini=%s, @ffin=%s, 
                        @idclin=%s, @fechapago=%s
                """, [
                    ide, cini, cfin,
                    fecha1, fecha2,
                    rini, rfin,
                    fini, ffin,
                    idclin, fechapago
                ])
                columns = [col[0] for col in cursor.description]
                raw_results = cursor.fetchall()

            results = [dict(zip(columns, row)) for row in raw_results]

            # Solo filas con algún saldo
            results = [
                r for r in results
                if any(r.get(col, 0) for col in ["total", "saldo", "saldo1", "dive"])
            ]

            # Subtotales por cliente
            subtotales = defaultdict(lambda: {col: 0 for col in ["total","saldo","saldo1"]})
            for r in results:
                cliente = r["idcliente"]
                for col in subtotales[cliente]:
                    subtotales[cliente][col] += r.get(col, 0)

            enriched = []
            for i, r in enumerate(results):
                enriched.append(r)
                next_cliente = results[i+1]["idcliente"] if i+1 < len(results) else None
                if r["idcliente"] != next_cliente:
                    enriched.append({
                        "subtotal": True,
                        "idcliente": r["idcliente"],
                        "nombre": r["nombre"],
                        **subtotales[r["idcliente"]]
                    })
            results = enriched


    context = {
        "title": "Reporte de Linea de Alimentos",
        "fecha": datetime.now(),
        "columns": columns,
        "results": results,
        "cini": cini,
        "cfin": cfin,
        "fecha1": fecha1,
        "fecha2": fecha2,
        "fechapago": fechapago
    }

    pdf = render_to_pdf("reporte_linea_alimentos_pdf.html", context)
    if pdf:
        return pdf
    return HttpResponse("Error al generar el PDF", status=500)


@login_required
def reporte_estado_general_pdf(request):
    ide = request.user.ide


    if "reporte" in request.GET:
        reporte = request.GET.get("reporte")

        if reporte == "estado_general":
            cini = request.GET.get('cini')
            cfin = request.GET.get('cfin')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))
            rini = request.GET.get('rini')
            rfin = request.GET.get('rfin')
            fini = request.GET.get('fini')
            ffin = request.GET.get('ffin')

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_ctes_edoctagral  
                        @ide=%s, @cini=%s, @cfin=%s, 
                        @fecha1=%s, @fecha2=%s, 
                        @rini=%s, @rfin=%s, 
                        @fini=%s, @ffin=%s
                """, [
                    ide, cini, cfin,
                    fecha1, fecha2,
                    rini, rfin,
                    fini, ffin
                ])

                # --- Primer resultset: clientes
                columns1 = [col[0] for col in cursor.description]
                results1 = [dict(zip(columns1, row)) for row in cursor.fetchall()]

                # --- Segundo resultset (detalle facturas + notas)
                results2 = []
                columns2 = []
                if cursor.nextset():
                    columns2 = [col[0] for col in cursor.description]
                    results2 = [dict(zip(columns2, row)) for row in cursor.fetchall()]

            # --- Calcular totales por cliente
            totales = defaultdict(lambda: {"cargo": 0, "abono": 0, "saldo": 0})
            for row in results2:
                cid = row["idcliente"]
                totales[cid]["cargo"] += row.get("cargo", 0) or 0
                totales[cid]["abono"] += row.get("abono", 0) or 0
                totales[cid]["saldo"] += row.get("saldo", 0) or 0


    context = {
        "title": "Reporte Estado General",
        "fecha": datetime.now(),
        "results1": results1,
        "results2": results2,
        "columns1": columns1,
        "columns2": columns2,
        "cini": cini,
        "cfin": cfin,
        "fini": fini,
        "ffin": ffin,
        "totales": totales,  # Aquí mandamos los totales
    }

    pdf = render_to_pdf("reporte_estado_general_pdf.html", context)
    if pdf:
        return pdf
    return HttpResponse("Error al generar el PDF", status=500)

@login_required
def reporte_recuperacion_cartera_pdf(request):
    ide = request.user.ide
    results, columns = [], []

    if "reporte" in request.GET:
        reporte = request.GET.get("reporte")

        if reporte == "recuperacion_cartera":
            cini = request.GET.get('cini')
            cfin = request.GET.get('cfin')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))
            rini = request.GET.get('rini')
            rfin = request.GET.get('rfin')
            fini = request.GET.get('fini')
            ffin = request.GET.get('ffin')
            fechapago = parse_fecha(request.GET.get('fechapago'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_ctes_recuperacion  
                        @ide=%s, @cini=%s, @cfin=%s, 
                        @fecha1=%s, @fecha2=%s, 
                        @rini=%s, @rfin=%s, 
                        @fini=%s, @ffin=%s, 
                        @fechapago=%s
                """, [
                    ide, cini, cfin,
                    fecha1, fecha2,
                    rini, rfin,
                    fini, ffin,
                    fechapago
                ])
                columns = [col[0] for col in cursor.description]
                raw_results = cursor.fetchall()

            results = [dict(zip(columns, row)) for row in raw_results]

            # Solo filas con algún saldo
            results = [
                r for r in results
                if any(r.get(col, 0) for col in ["total","pagado","saldo1","saldo2","saldo3","saldo4","saldo5","saldo6","saldo"])
            ]

            # Subtotales por cliente
            subtotales = defaultdict(lambda: {col: 0 for col in ["total","pagado","saldo1","saldo2","saldo3","saldo4","saldo5","saldo6","saldo"]})
            for r in results:
                cliente = r["idcliente"]
                for col in subtotales[cliente]:
                    subtotales[cliente][col] += r.get(col, 0)

            enriched = []
            for i, r in enumerate(results):
                enriched.append(r)
                next_cliente = results[i+1]["idcliente"] if i+1 < len(results) else None
                if r["idcliente"] != next_cliente:
                    enriched.append({
                        "subtotal": True,
                        "idcliente": r["idcliente"],
                        "nombre": r["nombre"],
                        **subtotales[r["idcliente"]]
                    })
            results = enriched


    context = {
        "title": "Reporte de Clientes",
        "fecha": datetime.now(),
        "columns": columns,
        "results": results,
        "cini": cini,
        "cfin": cfin,
        "fecha1": fecha1,
        "fecha2": fecha2,
        "fechapago": fechapago
    }

    pdf = render_to_pdf("reporte_recuperacion_cartera_pdf.html", context)
    if pdf:
        return pdf
    return HttpResponse("Error al generar el PDF", status=500)


@login_required
def reporte_ventas_devs_cuotas_pdf(request): 
    ide = request.user.ide
    results = []
    columns = []
    fecha1 = None
    fecha2 = None
    total_neto = 0
    total_cuota = 0
    total_porcentaje = 0

    if "reporte" in request.GET:
        reporte = request.GET.get("reporte")

        if reporte == "ventas":
            venini = request.GET.get('venini', '001')
            venfin = request.GET.get('venfin', '999')
            linini = request.GET.get('linini', '00101')
            linfin = request.GET.get('linfin', '99999')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_gralvtas_dev_ven
                        @ide=%s, @venini=%s, @venfin=%s, @linini=%s, @linfin=%s, 
                        @fecha1=%s, @fecha2=%s
                """, [ide, venini, venfin, linini, linfin, fecha1, fecha2])

                columns = [col[0] for col in cursor.description]
                raw_results = cursor.fetchall()

                results = [dict(zip(columns, row)) for row in raw_results]

    context = {
        "title": "Reporte de Ventas, Devs y Cuotas de Vendedores",
        "fecha": datetime.now(),
        'results': results,
        "fecha1": fecha1,
        "fecha2": fecha2,
        "total_neto": total_neto,
        "total_cuota": total_cuota,
        "total_porcentaje": total_porcentaje,
    }

    pdf = render_to_pdf("reporte_ventas_devs_cuotas_pdf.html", context)
    if pdf:
        return pdf

    return HttpResponse("Error al generar el PDF", status=500)

@login_required
def reporte_ventas_devs_ventas_pdf(request): 
    ide = request.user.ide
    results = []
    columns = []
    fecha1 = None
    fecha2 = None
    total_neto = 0
    total_cuota = 0
    total_porcentaje = 0

    if "reporte" in request.GET:
        reporte = request.GET.get("reporte")

        if reporte == "resumen":
            venini = request.GET.get('venini')
            venfin = request.GET.get('venfin', '999')
            linini = request.GET.get('linini')
            linfin = request.GET.get('linfin', '99999')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_r_gralvtas_utils_dev_ven
                        @ide=%s, @venini=%s, @venfin=%s, @linini=%s, @linfin=%s, 
                        @fecha1=%s, @fecha2=%s
                """, [ide, venini, venfin, linini, linfin, fecha1, fecha2])

                columns = [col[0] for col in cursor.description]
                raw_results = cursor.fetchall()

                results = [dict(zip(columns, row)) for row in raw_results]

    context = {
        "title": "Resumen de Ventas, Devs y Utilidades de Vendedores",
        "fecha": datetime.now(),
        'results': results,
        "fecha1": fecha1,
        "fecha2": fecha2,
        "total_neto": total_neto,
        "total_cuota": total_cuota,
        "total_porcentaje": total_porcentaje,
    }

    pdf = render_to_pdf("reporte_ventas_devs_ventas_pdf.html", context)
    if pdf:
        return pdf

    return HttpResponse("Error al generar el PDF", status=500)

def reporte_80_20_pdf(request):
    ide = request.user.ide
    results = []
    columns = []

    if "reporte" in request.GET:
        reporte = request.GET.get("reporte")

        if reporte == "8020":
            grp8020 = request.GET.get('grp8020')
            cini = request.GET.get('cini')
            cfin = request.GET.get('cfin')
            fecha1 = parse_fecha(request.GET.get('fecha1'))
            fecha2 = parse_fecha(request.GET.get('fecha2'))
            rini = request.GET.get('rini')
            rfin = request.GET.get('rfin')
            fini = request.GET.get('fini')
            ffin = request.GET.get('ffin')
            fechapago = parse_fecha(request.GET.get('fechapago'))

            with connection.cursor() as cursor:
                cursor.execute("""
                    EXEC dj_info_vtas_80_20_cte 
                        @grp8020=%s, @ide=%s, @cini=%s, @cfin=%s, 
                        @fecha1=%s, @fecha2=%s, 
                        @rini=%s, @rfin=%s, 
                        @fini=%s, @ffin=%s, 
                        @fechapago=%s
                """, [
                    grp8020, ide, cini, cfin,
                    fecha1, fecha2,
                    rini, rfin,
                    fini, ffin,
                    fechapago
                ])
                if cursor.description:   # ✅ evitar error si no hay SELECT
                    columns = [col[0] for col in cursor.description]
                    raw_results = cursor.fetchall()
                    results = [dict(zip(columns, row)) for row in raw_results]
                else:
                    columns = []
                    results = []

    context = {
        "title": "Reporte 80/20",
        "fecha": datetime.now(),
        "columns": columns,
        "results": results,
        "cini": cini,
        "cfin": cfin,
        "fecha1": fecha1,
        "fecha2": fecha2, 
        "fechapago": fechapago
    }

    pdf = render_to_pdf("reporte_80_20_pdf.html", context)
    if pdf:
        return pdf
    return HttpResponse("Error al generar el PDF", status=500)

def reporte_80_20_excel(request):
    ide = request.user.ide
    grp8020 = request.GET.get('grp8020')
    cini = request.GET.get('cini')
    cfin = request.GET.get('cfin')
    fecha1 = request.GET.get('fecha1')
    fecha2 = request.GET.get('fecha2')
    rini = request.GET.get('rini')
    rfin = request.GET.get('rfin')
    fini = request.GET.get('fini')
    ffin = request.GET.get('ffin')
    fechapago = request.GET.get('fechapago')

    # Ejecutar procedimiento almacenado
    with connection.cursor() as cursor:
        cursor.execute("""
            EXEC dj_info_vtas_80_20_cte 
                @grp8020=%s, @ide=%s, @cini=%s, @cfin=%s, 
                @fecha1=%s, @fecha2=%s, 
                @rini=%s, @rfin=%s, 
                @fini=%s, @ffin=%s, 
                @fechapago=%s
        """, [
            grp8020, ide, cini, cfin,
            fecha1, fecha2,
            rini, rfin,
            fini, ffin,
            fechapago
        ])
        columns = [col[0] for col in cursor.description]
        results = cursor.fetchall()

    # Crear libro Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte 80_20"

    # Encabezados
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = openpyxl.styles.Font(bold=True)

    # Datos
    for row_num, row_data in enumerate(results, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = cell_value

    # Ajuste de ancho
    for i, col in enumerate(columns, 1):
        max_length = max(
            (len(str(cell.value)) for cell in ws[get_column_letter(i)] if cell.value),
            default=10
        )
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    # Nombre dinámico del archivo
    nombre_archivo = f"Reporte_80_20_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response



@login_required
def add_cliente(request):
    ide = request.user.ide

    # --- BLOQUE AJAX (DEBE IR AL INICIO Y ESTAR BIEN IDENTADO) ---
    if request.GET.get('ajax') == '1':
        idvend = request.GET.get('idvend')

        try:
            rutas = utils.get_rutas_cliente(ide, idvend)
            return JsonResponse({'rutas': [r[0] for r in rutas]})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    # --- Despliegue de catálogos ---
    try:
        bancos = utils.get_banco_clientes()
    except Exception as e:
        messages.error(request, f"Error al cargar los bancos: {str(e)}")

    try:
        msg = utils.get_mensajes_clientes(ide)
    except Exception as e:
        messages.error(request, f"Error al cargar los mensajes: {str(e)}")

    try:
        vend = utils.get_vendedores_clientes(ide)
    except Exception as e:
        messages.error(request, f"Error al cargar los vendedores: {str(e)}")

    try:
        estado = utils.get_estados_cliente()
    except Exception as e:
        messages.error(request, f"Error al cargar los estados: {str(e)}")

    try:
        tipoc = utils.get_tipo_cliente()
    except Exception as e:
        messages.error(request, f"Error al cargar los tipos de cliente: {str(e)}")

    # --- Procesamiento del formulario POST ---
    if request.method == 'POST':
        ide = request.user.ide
        idcliente = request.POST.get('idcliente')
        nom_sat = request.POST.get('nom_sat')
        rfc_sat = request.POST.get('rfc')
        cp = request.POST.get('cp')
        ieps = request.POST.get('ieps')
        uso_cfdi = request.POST.get('uso_cfdi')
        rg_cfdi = request.POST.get('rg_cfdi')
        fp_cfdi = request.POST.get('fp_cfdi')
        mp_cfdi = request.POST.get('mp_cfdi')
        tnombancoordext = request.POST.get('tnombancoordext')
        trfcemisorctaord = request.POST.get('trfcemisorctaord')
        tctaordenante = request.POST.get('tctaordenante')
        sat_colonia = request.POST.get('sat_colonia')
        sat_mpio = request.POST.get('sat_mpio')
        sat_loc = request.POST.get('sat_loc')
        sat_edo = request.POST.get('sat_edo')
        distancia = request.POST.get('distancia')
        domicilio = request.POST.get('domicilio')
        exterior = request.POST.get('exterior')
        interior = request.POST.get('interior')
        colonia = request.POST.get('colonia')
        saldo = request.POST.get('saldo')
        credito = request.POST.get('credito')
        cxc = request.POST.get('cxc')
        cot = request.POST.get('cot')
        ped = request.POST.get('ped')
        fac = request.POST.get('fac')
        poblacion = request.POST.get('poblacion')
        municipio = request.POST.get('municipio')
        estado = request.POST.get('estado')
        idmc = request.POST.get('idmc')
        msg_post = request.POST.get('msg')  # para evitar sobrescribir `msg` del catálogo
        idvend = request.POST.get('idvend')
        ruta = request.POST.get('ruta')
        idtc = request.POST.get('idtc')
        cedula = request.POST.get('cedula')
        resp = request.POST.get('resp')
        aviso = request.POST.get('aviso')
        ref = request.POST.get('ref')
        

        with connection.cursor() as cursor:
            sql = 'EXEC dj_a_cliente ' + ', '.join(['%s'] * 40)
            cursor.execute(sql, [
                ide, idcliente, nom_sat, rfc_sat, cp, ieps, uso_cfdi, rg_cfdi, fp_cfdi, mp_cfdi,
                tnombancoordext, trfcemisorctaord, tctaordenante,
                sat_colonia, sat_mpio, sat_loc, sat_edo, distancia,
                domicilio, exterior, interior, colonia, saldo, credito,
                cxc, cot, ped, fac, poblacion, municipio, estado, idmc, msg_post, idvend,
                ruta, idtc, cedula, resp, aviso, ref
            ])
            result = cursor.fetchone()
        
        if result and result[0] == -1:
            messages.error(request, "El cliente ya existe.")
        else:
            messages.success(request, f"Cliente agregado correctamente. ID generado: {result[0]}")

        # Guardar correos
        try:
            utils.post_correo_cliente(request)
        except Exception as e:
            messages.error(request, f"Error al guardar los correos: {str(e)}")

        # Guardar teléfonos
        try:
            utils.post_telefono_cliente(request)
        except Exception as e:
            messages.error(request, f"Error al guardar los teléfonos: {str(e)}")

        # Guarda embarques
        try:
            utils.post_embarque_cliente(request)
        except Exception as e:
            messages.error(request, f'Error al agregar el embarque: {str(e)}')

        # Guarda domicilios
        try:
            utils.post_domicilio_cliente(request)
        except Exception as e:
            messages.error(request, f'Error al agregar el domicilio: {str(e)}')

        messages.success(request, "Cliente agregado correctamente.")
        return redirect('list_cliente')

    return render(request, 'cliente_form.html', {
        'bancos': bancos,
        'msg': msg,
        'vend': vend,
        'estado': estado,
        'tipoc': tipoc
    })

@login_required
def view_cliente(request, idcliente):
    ide = request.user.ide
    correos = []
    cliente = None
    clien = []
    telefono = []

    # --- BLOQUE AJAX ---
    if request.GET.get('ajax') == '1':
        idvend = request.GET.get('idvend')
        try:
            rutas = utils.get_rutas_cliente(ide, idvend)
            return JsonResponse({'rutas': [r[0] for r in rutas]})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    # --- Cargar catálogos ---
    try:
        bancos = utils.get_banco_clientes()
    except Exception as e:
        messages.error(request, f"Error al cargar los bancos: {str(e)}")

    try:
        msg = utils.get_mensajes_clientes(ide)
    except Exception as e:
        messages.error(request, f"Error al cargar los mensajes: {str(e)}")

    try:
        vend = utils.get_vendedores_clientes(ide)
    except Exception as e:
        messages.error(request, f"Error al cargar los vendedores: {str(e)}")

    try:
        estado = utils.get_estados_cliente()
    except Exception as e:
        messages.error(request, f"Error al cargar los estados: {str(e)}")

    try:
        tipoc = utils.get_tipo_cliente()
    except Exception as e:
        messages.error(request, f"Error al cargar los tipos de cliente: {str(e)}")

    if request.method == 'POST':
        # Campos del formulario
        idx = request.POST.get('idx')
        nom_sat = request.POST.get('nom_sat')
        rfc_sat = request.POST.get('rfc')
        cp = request.POST.get('cp')
        ieps = request.POST.get('ieps')
        uso_cfdi = request.POST.get('uso_cfdi')
        rg_cfdi = request.POST.get('rg_cfdi')
        fp_cfdi = request.POST.get('fp_cfdi')
        mp_cfdi = request.POST.get('mp_cfdi')
        tnombancoordext = request.POST.get('tnombancoordext')
        trfcemisorctaord = request.POST.get('trfcemisorctaord')
        tctaordenante = request.POST.get('tctaordenante')
        sat_colonia = request.POST.get('sat_colonia')
        sat_mpio = request.POST.get('sat_mpio')
        sat_loc = request.POST.get('sat_loc')
        sat_edo = request.POST.get('sat_edo')
        distancia = request.POST.get('distancia')
        domicilio = request.POST.get('domicilio')
        exterior = request.POST.get('exterior')
        interior = request.POST.get('interior')
        colonia = request.POST.get('colonia')
        saldo = request.POST.get('saldo')
        credito = request.POST.get('credito')
        cxc = request.POST.get('cxc')
        cot = request.POST.get('cot')
        ped = request.POST.get('ped')
        fac = request.POST.get('fac')
        poblacion = request.POST.get('poblacion')
        municipio = request.POST.get('municipio')
        estado = request.POST.get('estado')
        idmc = request.POST.get('idmc')
        msg_post = request.POST.get('msg')  # para evitar sobrescribir `msg` del catálogo
        idvend = request.POST.get('idvend')
        ruta = request.POST.get('ruta')
        idtc = request.POST.get('idtc')
        cedula = request.POST.get('cedula')
        resp = request.POST.get('resp')
        aviso = request.POST.get('aviso')
        ref = request.POST.get('ref')

        # Guardar cliente
        try:
            with connection.cursor() as cursor:
                placeholders = ', '.join(['%s'] * 39)
                sql = f'EXEC dj_e_cliente {placeholders}'
                cursor.execute(sql, [
                    idx, nom_sat, rfc_sat, cp, ieps, uso_cfdi, rg_cfdi, fp_cfdi, mp_cfdi,
                    tnombancoordext, trfcemisorctaord, tctaordenante, sat_colonia, sat_mpio,
                    sat_loc, sat_edo, distancia, domicilio, exterior, interior, colonia,
                    saldo, credito, cxc, cot, ped, fac, poblacion, municipio, estado,
                    idmc, msg_post, idvend, ruta, idtc, cedula, resp, aviso, ref
                ])
        except Exception as e:
            messages.error(request, f"Error al guardar el cliente: {str(e)}")

        # Guardar correos
        try:
            utils.post_correo_cliente(request)
        except Exception as e:
            messages.error(request, f"Error al guardar los correos: {str(e)}")

        # Guardar teléfonos
        try:
            utils.post_telefono_cliente(request)
        except Exception as e:
            messages.error(request, f"Error al guardar los teléfonos: {str(e)}")

        # Guarda embarques
        try:
            utils.post_embarque_cliente(request)
        except Exception as e:
            messages.error(request, f'Error al agregar el embarque: {str(e)}')

        # Guarda domicilios
        try:
            utils.post_domicilio_cliente(request)
        except Exception as e:
            messages.error(request, f'Error al agregar el domicilio: {str(e)}')

        messages.success(request, "Cliente editado correctamente.")
        return redirect('list_cliente')

    # --- GET: cargar datos del cliente y listas ---
    try:
        cliente = utils.get_clientes(idcliente, ide)
    except Exception as e:
        messages.error(request, f"Error al obtener datos del cliente: {str(e)}")

    try:
        clien = utils.get_clientes(idcliente, ide)
    except Exception as e:
        messages.error(request, f"Error al cargar los clientes: {str(e)}")
        clien = []

    try:
        correos = utils.get_correo_cliente(idcliente, ide)
    except Exception as e:
        messages.error(request, f'Error al cargar los correos: {str(e)}')

    try:
        telefono = utils.get_telefono_cliente(idcliente, ide)
    except Exception as e:
        messages.error(request, f'Error al cargar los teléfonos: {str(e)}')
    
    # embarques 
    try:
        embar = utils.get_lista_embarques(ide)
    except Exception as e:
        messages.error(request, f'Error al cargar los embarques: {str(e)}')
    
    try:
        embarques = utils.get_datos_embarques(idcliente, ide)
    except Exception as e:
        messages.error(request, f'Error al cargar los datos de los embarques: {str(e)}')

    # domicilio
    try:
        domicilios = utils.get_lista_domicilio(idcliente, ide)
    except Exception as e:
        messages.error(request, f'Error al cargar los datos de los embarques: {str(e)}')
    

    return render(request, 'cliente_view.html', {
        'clientes': cliente,
        'bancos': bancos,
        'msg': msg,
        'vend': vend,
        'clien': clien,
        'estado': estado,
        'tipoc': tipoc,
        'correos': correos,
        'telefonos': telefono,
        'embar': embar,
        'embarques': embarques,
        'domicilio': domicilios
    })

    
@login_required
def edit_cliente(request, idcliente):
    ide = request.user.ide
    correos = []
    cliente = None
    clien = []
    telefono = []


    # --- BLOQUE AJAX ---
    if request.GET.get('ajax') == '1':
        idvend = request.GET.get('idvend')
        try:
            rutas = utils.get_rutas_cliente(ide, idvend)
            return JsonResponse({'rutas': [r[0] for r in rutas]})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    # --- Cargar catálogos ---
    try:
        bancos = utils.get_banco_clientes()
    except Exception as e:
        messages.error(request, f"Error al cargar los bancos: {str(e)}")

    try:
        msg = utils.get_mensajes_clientes(ide)
    except Exception as e:
        messages.error(request, f"Error al cargar los mensajes: {str(e)}")

    try:
        vend = utils.get_vendedores_clientes(ide)
    except Exception as e:
        messages.error(request, f"Error al cargar los vendedores: {str(e)}")

    try:
        estado = utils.get_estados_cliente()
    except Exception as e:
        messages.error(request, f"Error al cargar los estados: {str(e)}")

    try:
        tipoc = utils.get_tipo_cliente()
    except Exception as e:
        messages.error(request, f"Error al cargar los tipos de cliente: {str(e)}")

    if request.method == 'POST':
        # Campos del formulario
        idx = request.POST.get('idx')
        nom_sat = request.POST.get('nom_sat')
        rfc_sat = request.POST.get('rfc')
        cp = request.POST.get('cp')
        ieps = request.POST.get('ieps')
        uso_cfdi = request.POST.get('uso_cfdi')
        rg_cfdi = request.POST.get('rg_cfdi')
        fp_cfdi = request.POST.get('fp_cfdi')
        mp_cfdi = request.POST.get('mp_cfdi')
        tnombancoordext = request.POST.get('tnombancoordext')
        trfcemisorctaord = request.POST.get('trfcemisorctaord')
        tctaordenante = request.POST.get('tctaordenante')
        sat_colonia = request.POST.get('sat_colonia')
        sat_mpio = request.POST.get('sat_mpio')
        sat_loc = request.POST.get('sat_loc')
        sat_edo = request.POST.get('sat_edo')
        distancia = request.POST.get('distancia')
        domicilio = request.POST.get('domicilio')
        exterior = request.POST.get('exterior')
        interior = request.POST.get('interior')
        colonia = request.POST.get('colonia')
        saldo = request.POST.get('saldo')
        credito = request.POST.get('credito')
        cxc = request.POST.get('cxc')
        cot = request.POST.get('cot')
        ped = request.POST.get('ped')
        fac = request.POST.get('fac')
        poblacion = request.POST.get('poblacion')
        municipio = request.POST.get('municipio')
        estado = request.POST.get('estado')
        idmc = request.POST.get('idmc')
        msg_post = request.POST.get('msg')  # para evitar sobrescribir `msg` del catálogo
        idvend = request.POST.get('idvend')
        ruta = request.POST.get('ruta')
        idtc = request.POST.get('idtc')
        cedula = request.POST.get('cedula')
        resp = request.POST.get('resp')
        aviso = request.POST.get('aviso')
        ref = request.POST.get('ref')

        # Guardar cliente
        try:
            with connection.cursor() as cursor:
                placeholders = ', '.join(['%s'] * 39)
                sql = f'EXEC dj_e_cliente {placeholders}'
                cursor.execute(sql, [
                    idx, nom_sat, rfc_sat, cp, ieps, uso_cfdi, rg_cfdi, fp_cfdi, mp_cfdi,
                    tnombancoordext, trfcemisorctaord, tctaordenante, sat_colonia, sat_mpio,
                    sat_loc, sat_edo, distancia, domicilio, exterior, interior, colonia,
                    saldo, credito, cxc, cot, ped, fac, poblacion, municipio, estado,
                    idmc, msg_post, idvend, ruta, idtc, cedula, resp, aviso, ref
                ])
        except Exception as e:
            messages.error(request, f"Error al guardar el cliente: {str(e)}")

        # Guardar correos
        try:
            utils.post_correo_cliente(request)
        except Exception as e:
            messages.error(request, f"Error al guardar los correos: {str(e)}")

        # Guardar teléfonos
        try:
            utils.post_telefono_cliente(request)
        except Exception as e:
            messages.error(request, f"Error al guardar los teléfonos: {str(e)}")

        # Guarda embarques
        try:
            utils.post_embarque_cliente(request)
        except Exception as e:
            messages.error(request, f'Error al agregar el embarque: {str(e)}')

        # Guarda domicilios
        try:
            utils.post_domicilio_cliente(request)
        except Exception as e:
            messages.error(request, f'Error al agregar el domicilio: {str(e)}')

        messages.success(request, "Cliente editado correctamente.")
        return redirect('list_cliente')

    # --- GET: cargar datos del cliente y listas ---
    try:
        cliente = utils.get_clientes(idcliente, ide)
    except Exception as e:
        messages.error(request, f"Error al obtener datos del cliente: {str(e)}")

    try:
        clien = utils.get_clientes(idcliente, ide)
    except Exception as e:
        messages.error(request, f"Error al cargar los clientes: {str(e)}")
        clien = []

    try:
        correos = utils.get_correo_cliente(idcliente, ide)
    except Exception as e:
        messages.error(request, f'Error al cargar los correos: {str(e)}')

    try:
        telefono = utils.get_telefono_cliente(idcliente, ide)
    except Exception as e:
        messages.error(request, f'Error al cargar los teléfonos: {str(e)}')
    
    # embarques 
    try:
        embar = utils.get_lista_embarques(ide)
    except Exception as e:
        messages.error(request, f'Error al cargar los embarques: {str(e)}')
    
    try:
        embarques = utils.get_datos_embarques(idcliente, ide)
    except Exception as e:
        messages.error(request, f'Error al cargar los datos de los embarques: {str(e)}')

    # domicilio
    try:
        domicilios = utils.get_lista_domicilio(idcliente, ide)
    except Exception as e:
        messages.error(request, f'Error al cargar los datos de los embarques: {str(e)}')
    

    return render(request, 'cliente_form.html', {
        'clientes': cliente,
        'bancos': bancos,
        'msg': msg,
        'vend': vend,
        'clien': clien,
        'estado': estado,
        'tipoc': tipoc,
        'correos': correos,
        'telefonos': telefono,
        'embar': embar,
        'embarques': embarques,
        'domicilio': domicilios
    })

@login_required
def delete_cliente(request, idx):
    
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_d_cliente %s', [idx])
        
        messages.error(request, 'Eliminado correctamente')
    return redirect('list_cliente')

# Funcionalidad de clientes con correo

@login_required
def list_cliente_correo(request):
    idcliente = request.user.ide
    
    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_l_ubicacion %s', [idcliente])
        ubicacion = cursor.fetchall()
    
    return render(request, 'ubicacion.html', {'ubicaciones': ubicacion})

@login_required
def edit_cliente_correo(request, idx):
   
    if request.method == 'POST':
        idubica = request.POST.get('idubica')  # <- Sin la coma

        with connection.cursor() as cursor:
            cursor.execute('EXEC dj_e_ubicacion %s, %s', [ idubica, idx])
        
        messages.success(request, 'Editado correctamente')
        return redirect('list_ubicacion')
    
    else:
        with connection.cursor() as cursor:
                cursor.execute("EXEC dj_l_e_ubicacion %s", [idx])
                row = cursor.fetchone()
        if row:
                ubicacion = {
                    # Asegúrate de que los índices son correctos
                    'idubica': row[2]
                }
    return render(request, 'ubicacion_form.html', {'ubicaciones': ubicacion})

# Funcionalidad de almacen 
def list_almacen(request):
    ide = request.user.ide

    with connection.cursor() as cursor:
        cursor.execute('EXEC dj_l_prod_alma %s', [ide])
        almacen = cursor.fetchall()
    return render(request, 'almacen.html', {'almacenes': almacen})
